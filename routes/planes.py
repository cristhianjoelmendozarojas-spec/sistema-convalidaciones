# routes/planes.py
from flask import (
    Blueprint,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    jsonify,
    send_file,
)
from routes.auth import modulo_requerido
from db.conexion import get_connection
import io
import os

bp_planes = Blueprint("planes", __name__)

CICLOS_VALIDOS = {"I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"}
ORDEN_CICLOS = ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X"]

# ── Helpers ───────────────────────────────────────────────────────


def _leer_archivo(archivo, ext):
    if ext in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(archivo, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, None
        return [str(c).strip().upper() if c else "" for c in rows[0]], rows[1:]
    else:
        import csv
        import io as _io

        contenido = archivo.read().decode("utf-8-sig")
        reader = csv.reader(_io.StringIO(contenido))
        rows = list(reader)
        if not rows:
            return None, None
        return [c.strip().upper() for c in rows[0]], rows[1:]


def _validar_fila(fila_dict):
    nombre = str(fila_dict.get("NOMBRE DEL CURSO") or "").strip()
    ciclo = str(fila_dict.get("CICLO") or "").strip().upper()
    cred = fila_dict.get("CREDITOS")
    if "TOTAL" in nombre.upper() or "TOTAL" in ciclo:
        return False, True, []
    errores = []
    if ciclo not in CICLOS_VALIDOS:
        errores.append(f'Ciclo "{ciclo}" no válido')
    if not nombre:
        errores.append("Nombre del curso vacío")
    try:
        c = int(str(cred).strip()) if cred else None
        if c is None or c <= 0:
            errores.append("Créditos debe ser entero positivo")
    except Exception:
        errores.append("Créditos no es número")
    return len(errores) == 0, False, errores


# ── Rutas ─────────────────────────────────────────────────────────


@bp_planes.route("/")
@modulo_requerido("planes")
def index():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT p.id, p.nombre_plan, p.tipo_plan, p.periodo_academico, p.fecha_importacion,
               COUNT(c.id) as total_cursos, COALESCE(SUM(c.creditos),0) as total_creditos
        FROM planes_estudio p
        LEFT JOIN cursos_plan c ON c.plan_id = p.id
        GROUP BY p.id ORDER BY p.nombre_plan, p.periodo_academico
    """)
    planes = cur.fetchall()

    # Obtener nombres únicos para el combobox en importar
    cur.execute("SELECT DISTINCT nombre_plan FROM planes_estudio ORDER BY nombre_plan")
    nombres_existentes = [r["nombre_plan"] for r in cur.fetchall()]
    cur.close()
    conn.close()

    # Agrupar por nombre_plan → {nombre: {tipo, periodos: [...]}}
    grupos = {}
    for p in planes:
        n = p["nombre_plan"]
        if n not in grupos:
            grupos[n] = {"nombre": n, "tipo_plan": p["tipo_plan"], "periodos": []}
        grupos[n]["periodos"].append(p)

    locales = {n: g for n, g in grupos.items() if g["tipo_plan"] == "local"}
    externos = {n: g for n, g in grupos.items() if g["tipo_plan"] == "externo"}

    return render_template(
        "planes/lista.html",
        locales=locales,
        externos=externos,
        nombres_existentes=nombres_existentes,
    )


@bp_planes.route("/importar", methods=["GET", "POST"])
@modulo_requerido("planes")
def importar():
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT nombre_plan FROM planes_estudio ORDER BY nombre_plan")
    nombres_existentes = [r["nombre_plan"] for r in cur.fetchall()]
    cur.close()
    conn.close()

    if request.method == "GET":
        return render_template(
            "planes/importar.html", nombres_existentes=nombres_existentes
        )

    nombre_plan = request.form.get("nombre_plan", "").strip()
    tipo_plan = request.form.get("tipo_plan", "local")
    periodo_academico = request.form.get("periodo_academico", "").strip()
    archivo = request.files.get("archivo")

    if not nombre_plan:
        flash("El nombre del plan es obligatorio.", "danger")
        return render_template(
            "planes/importar.html", nombres_existentes=nombres_existentes
        )
    if not archivo or not archivo.filename:
        flash("Selecciona un archivo.", "danger")
        return render_template(
            "planes/importar.html", nombres_existentes=nombres_existentes
        )

    ext = os.path.splitext(archivo.filename)[1].lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        flash("Formato no válido. Usa .xlsx, .xls o .csv", "danger")
        return render_template(
            "planes/importar.html", nombres_existentes=nombres_existentes
        )

    try:
        encabezados, datos = _leer_archivo(archivo, ext)
        if encabezados is None:
            flash("El archivo está vacío.", "danger")
            return render_template(
                "planes/importar.html", nombres_existentes=nombres_existentes
            )

        col_map = {}
        for col in ["CICLO", "CÓDIGO", "NOMBRE DEL CURSO", "CREDITOS", "PRERREQUISITO"]:
            try:
                col_map[col] = encabezados.index(col)
            except ValueError:
                pass

        if "NOMBRE DEL CURSO" not in col_map or "CICLO" not in col_map:
            flash("El archivo debe tener columnas CICLO y NOMBRE DEL CURSO.", "danger")
            return render_template(
                "planes/importar.html", nombres_existentes=nombres_existentes
            )

        filas_validas, errores_total = [], []
        for num, fila in enumerate(datos, 2):
            fila_dict = {
                col: (fila[idx] if idx < len(fila) else None)
                for col, idx in col_map.items()
            }
            es_valida, es_total, errs = _validar_fila(fila_dict)
            if es_total:
                continue
            if es_valida:
                filas_validas.append(fila_dict)
            else:
                errores_total.append(f"Fila {num}: {', '.join(errs)}")

        if not filas_validas:
            flash("No se encontraron cursos válidos en el archivo.", "danger")
            return render_template(
                "planes/importar.html", nombres_existentes=nombres_existentes
            )

        conn = get_connection()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO planes_estudio (nombre_plan,tipo_plan,periodo_academico) VALUES (%s,%s,%s)",
            (nombre_plan, tipo_plan, periodo_academico),
        )
        plan_id = cur.lastrowid
        if plan_id is None:
            raise Exception(
                "No se pudo obtener el ID del plan de estudio creado. Verificar secuencia de la base de datos."
            )

        for f in filas_validas:
            cur.execute(
                """INSERT INTO cursos_plan (plan_id,ciclo,codigo,nombre_curso,creditos,prerrequisito)
                           VALUES (%s,%s,%s,%s,%s,%s)""",
                (
                    plan_id,
                    str(f.get("CICLO") or "").strip().upper(),
                    str(f.get("CÓDIGO") or "").strip() or None,
                    str(f.get("NOMBRE DEL CURSO") or "").strip(),
                    int(str(f.get("CREDITOS") or 0).strip()),
                    str(f.get("PRERREQUISITO") or "").strip() or None,
                ),
            )

        conn.commit()
        msg = f'Plan "{nombre_plan}" · {periodo_academico} importado con {len(filas_validas)} curso(s).'
        if errores_total:
            msg += f" {len(errores_total)} fila(s) con errores omitida(s)."
        flash(msg, "success" if not errores_total else "warning")
        return redirect(url_for("planes.index"))

    except Exception as e:
        try:
            conn.rollback()
        except Exception:
            pass
        flash(f"Error al procesar el archivo: {str(e)}", "danger")
        return render_template(
            "planes/importar.html", nombres_existentes=nombres_existentes
        )
    finally:
        try:
            cur.close()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass


@bp_planes.route("/plantilla-excel")
def descargar_plantilla():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan de Estudios"
    borde = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for i, (t, w) in enumerate(
        zip(
            ["CICLO", "CÓDIGO", "NOMBRE DEL CURSO", "CREDITOS", "PRERREQUISITO"],
            [8, 14, 40, 12, 14],
        ),
        1,
    ):
        c = ws.cell(row=1, column=i, value=t)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor="0C1D3A")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = borde
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 28
    for ri, fila in enumerate(
        [
            ("I", "P02A1101", "Matemática I", 4, ""),
            ("I", "P02A1102", "Redacción y Comunicación", 4, ""),
            ("II", "P02A1107", "Matemática II", 4, "P02A1101"),
            ("III", "", "Electivo 1", 2, ""),
        ],
        2,
    ):
        for ci, v in enumerate(fila, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(color="888888", italic=True, size=9)
            c.border = borde
            c.alignment = Alignment(
                horizontal="center" if ci in (1, 4) else "left", vertical="center"
            )
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="plantilla_plan_estudios.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── API JSON ──────────────────────────────────────────────────────


@bp_planes.route("/periodos/<path:nombre_plan>")
def periodos(nombre_plan):
    """Devuelve todos los periodos de un plan agrupado por nombre."""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute(
        """
        SELECT p.id, p.periodo_academico, p.tipo_plan,
               COUNT(c.id) as total_cursos, COALESCE(SUM(c.creditos),0) as total_creditos
        FROM planes_estudio p
        LEFT JOIN cursos_plan c ON c.plan_id = p.id
        WHERE p.nombre_plan = %s
        GROUP BY p.id ORDER BY p.periodo_academico
    """,
        (nombre_plan,),
    )
    periodos = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify({"ok": True, "periodos": periodos})


@bp_planes.route("/cursos/<int:plan_id>")
def cursos(plan_id):
    """Devuelve cursos de un plan específico."""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM planes_estudio WHERE id=%s", (plan_id,))
    plan = cur.fetchone()
    if not plan:
        return jsonify({"ok": False}), 404
    _oc = "CASE ciclo WHEN 'I' THEN 1 WHEN 'II' THEN 2 WHEN 'III' THEN 3 WHEN 'IV' THEN 4 WHEN 'V' THEN 5 WHEN 'VI' THEN 6 WHEN 'VII' THEN 7 WHEN 'VIII' THEN 8 WHEN 'IX' THEN 9 WHEN 'X' THEN 10 END"
    cur.execute(
        f"""
        SELECT ciclo, codigo, nombre_curso, creditos, prerrequisito
        FROM cursos_plan WHERE plan_id=%s
        ORDER BY {_oc}, nombre_curso
    """,
        (plan_id,),
    )
    cursos = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(
        {
            "ok": True,
            "plan": plan["nombre_plan"],
            "periodo": plan["periodo_academico"],
            "cursos": cursos,
        }
    )


@bp_planes.route("/editar/<int:plan_id>", methods=["POST"])
def editar(plan_id):
    d = request.get_json()
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute(
            """UPDATE planes_estudio
                       SET nombre_plan=%s, tipo_plan=%s, periodo_academico=%s
                       WHERE id=%s""",
            (
                d.get("nombre_plan", "").strip(),
                d.get("tipo_plan", "local"),
                d.get("periodo_academico", "").strip(),
                plan_id,
            ),
        )
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@bp_planes.route("/eliminar/<int:plan_id>", methods=["POST"])
def eliminar(plan_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM planes_estudio WHERE id=%s", (plan_id,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@bp_planes.route("/eliminar-grupo/<nombre_plan>", methods=["POST"])
def eliminar_grupo(nombre_plan):
    """Elimina todos los periodos de un plan."""
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM planes_estudio WHERE nombre_plan=%s", (nombre_plan,))
        conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()


@bp_planes.route("/nombres")
def nombres():
    """API: lista de nombres únicos para el combobox."""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT DISTINCT nombre_plan FROM planes_estudio ORDER BY nombre_plan")
    data = [r["nombre_plan"] for r in cur.fetchall()]
    cur.close()
    conn.close()
    return jsonify(data)
