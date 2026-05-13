# routes/solicitudes.py — Schema v2
# Lógica de negocio movida a services/solicitud_service.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, send_file
from routes.logger import registrar
from routes.auth import modulo_requerido
from routes.generar_word import invalidar_cache
from services import solicitud_service as service
from services.postulantes_service import get_postulantes_lista
from psycopg2.extras import RealDictCursor
from db.conexion import get_connection

def generar_codigo_local(anio):
    """Versión local para evitar problemas de import."""
    from db.conexion import fetch_one
    try:
        row = fetch_one(
            "SELECT codigo FROM solicitudes WHERE codigo LIKE %s ORDER BY codigo DESC LIMIT 1",
            (f'SIMULACION-FCS-%-{anio}',)
        )
        if not row:
            return f"SIMULACION-FCS-001-{anio}"
        codigo = row.get('codigo') if isinstance(row, dict) else (row[0] if row else '')
        if not codigo:
            return f"SIMULACION-FCS-001-{anio}"
        return f"SIMULACION-FCS-{int(codigo.split('-')[2]) + 1:03d}-{anio}"
    except:
        return f"SIMULACION-FCS-001-{anio}"

import base64
import secrets
from datetime import datetime

bp = Blueprint('solicitudes', __name__)

CICLOS   = ['I','II','III','IV','V','VI','VII','VIII','IX','X']
PERIODOS = ['2025-1','2025-2','2026-1','2026-2','2027-1','2027-2','2028-1','2028-2']
_ORDEN_CICLO = "CASE ciclo WHEN 'I' THEN 1 WHEN 'II' THEN 2 WHEN 'III' THEN 3 WHEN 'IV' THEN 4 WHEN 'V' THEN 5 WHEN 'VI' THEN 6 WHEN 'VII' THEN 7 WHEN 'VIII' THEN 8 WHEN 'IX' THEN 9 WHEN 'X' THEN 10 END"

# Alias para compatibilidad
get_solicitud_completa = service.get_solicitud_completa
generar_codigo = generar_codigo_local

get_planes_por_tipo = service.get_planes_por_tipo
get_cursos_plan = service.get_cursos_plan
buscar_postulante = service.buscar_postulante
verificar_duplicado = service.verificar_duplicado

def _get_s_basico(id):
    """Obtiene datos básicos de la solicitud + postulante para las etapas."""
    from services.solicitud_service import get_solicitud_completa
    return get_solicitud_completa(id)


# ─────────────────────────────────────────────────────────────────
# LISTA
# ─────────────────────────────────────────────────────────────────

@bp.route('/')
@modulo_requerido('solicitudes')
def index():
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(50, max(10, int(request.args.get('per_page', 20))))
    offset = (page - 1) * per_page
    
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    
    cur.execute("SELECT COUNT(*) as total FROM solicitudes")
    total = cur.fetchone()['total']
    total_pages = (total + per_page - 1) // per_page
    
    cur.execute(f"""
        SELECT s.id, s.codigo, s.fecha_emision, s.estado, s.fecha_registro, s.confirmado,
               s.plan_externo_id, s.fecha_confirmacion, s.estado_confirmacion, s.observacion,
               COALESCE(p.apellidos_nombres,'Sin nombre') AS nombre,
               COALESCE(p.dni,'—')      AS dni,
               COALESCE(p.programa,'—') AS programa,
               COALESCE(p.correo,'') AS correo,
               COALESCE(p.celular,'') AS celular
        FROM solicitudes s
        LEFT JOIN postulantes p ON s.postulante_id = p.id
        ORDER BY s.fecha_registro DESC
        LIMIT %s OFFSET %s
    """, (per_page, offset))
    solicitudes = cur.fetchall()
    cur.close(); conn.close()
    return render_template('solicitudes/lista.html', solicitudes=solicitudes, 
                         page=page, per_page=per_page, total=total, 
                         total_pages=total_pages)


# ─────────────────────────────────────────────────────────────────
# VER DETALLE
# ─────────────────────────────────────────────────────────────────

@bp.route('/ver/<int:id>')
def ver(id):
    try:
        s = get_solicitud_completa(id)
        if not s:
            flash('Solicitud no encontrada.', 'danger')
            return redirect(url_for('solicitudes.index'))
        return render_template('solicitudes/detalle.html', s=s)
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Error al cargar detalle: {e}', 'danger')
        return redirect(url_for('solicitudes.index'))


# ─────────────────────────────────────────────────────────────────
# NUEVA SOLICITUD
# ─────────────────────────────────────────────────────────────────

@bp.route('/nueva', methods=['GET','POST'])
def nueva():
    if request.method == 'POST':
        from datetime import datetime as dt
        data  = request.form
        anio  = data.get('anio', str(dt.now().year))
        codigo = generar_codigo(anio)

        postulante_id = data.get('postulante_id') or None
        if postulante_id:
            try: postulante_id = int(postulante_id)
            except: postulante_id = None

        if not postulante_id:
            dni_val = data.get('dni','').strip()
            if dni_val:
                conn_p = get_connection(); cur_p = conn_p.cursor(dictionary=True)
                cur_p.execute("SELECT id FROM postulantes WHERE dni=%s LIMIT 1", (dni_val,))
                row_p = cur_p.fetchone()
                if row_p: postulante_id = row_p['id']
                cur_p.close(); conn_p.close()

        if postulante_id:
            conn_dup = get_connection(); cur_dup = conn_dup.cursor(dictionary=True)
            cur_dup.execute("SELECT id, codigo FROM solicitudes WHERE postulante_id=%s LIMIT 1",
                            (postulante_id,))
            dup = cur_dup.fetchone()
            cur_dup.close(); conn_dup.close()
            if dup:
                flash(f'Este postulante ya tiene la solicitud {dup["codigo"]} registrada.', 'warning')
                return redirect(url_for('solicitudes.nueva'))

        # Si viene carrera_id del formulario, actualizar sesión
        carrera_id = data.get('carrera_id')
        periodo_selected = data.get('periodo', '')
        if carrera_id:
            try:
                conn_c = get_connection(); cur_c = conn_c.cursor(dictionary=True)
                cur_c.execute("""
                    SELECT c.*, f.nombre AS facultad_nombre,
                           cp.costo_convalidacion, cp.costo_examen
                    FROM carreras c
                    JOIN facultades f ON c.facultad_id = f.id
                    LEFT JOIN carreras_periodos cp ON cp.carrera_id = c.id AND cp.periodo=%s
                    WHERE c.id=%s
                """, (periodo_selected, int(carrera_id)))
                carrera = cur_c.fetchone()
                if carrera:
                    session['carrera_id'] = carrera['id']
                    session['carrera_nombre'] = carrera['nombre']
                    session['facultad_nombre'] = carrera['facultad_nombre']
                    session['periodo'] = periodo_selected
                    session['costo_credito_carrera'] = float(carrera['costo_convalidacion'] or 60)
                    session['costo_examen_carrera'] = float(carrera['costo_examen'] or 130)
                cur_c.close(); conn_c.close()
            except: pass

        carrera_id = session.get('carrera_id')

        costo_cred = session.get('costo_credito_carrera', float(data.get('costo_credito', 60)))
        costo_exam = session.get('costo_examen_carrera',  float(data.get('costo_examen', 130)))

        conn = get_connection(); cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO solicitudes (codigo, postulante_id, carrera_id, fecha_emision, observacion, costo_credito, costo_examen)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (codigo, postulante_id, carrera_id,
                  data.get('fecha_emision') or dt.now().strftime('%Y-%m-%d'),
                  data.get('observacion',''),
                  costo_cred, costo_exam))
            conn.commit()
            nuevo_id = cur.lastrowid
            if not nuevo_id:
                flash(f'Solicitud creada — {codigo} (ID: {nuevo_id})', 'success')
                return redirect(url_for('solicitudes.index'))
            registrar('crear', 'solicitudes', f'Solicitud creada: {codigo}', nuevo_id)
            flash(f'Solicitud creada — {codigo}', 'success')
            return redirect(url_for('solicitudes.convalidar', id=nuevo_id))
        except Exception as e:
            conn.rollback(); flash(f'Error: {str(e)}', 'danger')
        finally:
            cur.close(); conn.close()

    from datetime import datetime as dt
    anio_actual    = dt.now().year
    codigo_preview = generar_codigo(anio_actual)
    costo_credito_default = session.get('costo_credito_carrera', 60)
    costo_examen_default  = session.get('costo_examen_carrera',  130)
    carrera_nombre        = session.get('carrera_nombre', '')
    carrera_id_session    = session.get('carrera_id')
    facultad_nombre       = session.get('facultad_nombre', '')

    conn = get_connection(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT c.id, c.nombre, cp.costo_convalidacion, cp.costo_examen,
               cp.periodo, f.nombre AS facultad_nombre
        FROM carreras c
        JOIN facultades f ON c.facultad_id = f.id
        JOIN carreras_periodos cp ON cp.carrera_id = c.id
        WHERE c.estado = 'activo'
        ORDER BY cp.periodo, f.nombre, c.nombre
    """)
    carreras = cur.fetchall()
    cur.execute("""
        SELECT DISTINCT cp.periodo FROM carreras_periodos cp
        JOIN carreras c ON cp.carrera_id = c.id
        WHERE c.estado='activo' AND cp.periodo IS NOT NULL AND cp.periodo != ''
        ORDER BY cp.periodo DESC
    """)
    periodos = [r['periodo'] for r in cur.fetchall()]
    cur.close(); conn.close()

    return render_template('solicitudes/formulario.html',
                           anio_actual=anio_actual, codigo_preview=codigo_preview,
                           costo_credito_default=costo_credito_default,
                           costo_examen_default=costo_examen_default,
                           carrera_nombre=carrera_nombre,
                           carrera_id_session=carrera_id_session,
                           facultad_nombre=facultad_nombre,
                           carreras=carreras, periodos=periodos)


# ─────────────────────────────────────────────────────────────────
# EDITAR SOLICITUD
# ─────────────────────────────────────────────────────────────────

@bp.route('/editar/<int:id>', methods=['GET','POST'])
def editar(id):
    s = _get_s_basico(id)
    if not s:
        flash('No encontrado.','danger')
        return redirect(url_for('solicitudes.index'))

    if request.method == 'POST':
        data = request.form
        conn = get_connection(); cur = conn.cursor()
        try:
            cur.execute("""
                UPDATE solicitudes
                SET fecha_emision=%s, observacion=%s,
                    costo_credito=%s, costo_examen=%s
                WHERE id=%s
            """, (data.get('fecha_emision'), data.get('observacion',''),
                  float(data.get('costo_credito',60)), float(data.get('costo_examen',130)), id))
            conn.commit()
            invalidar_cache(id)
            registrar('editar', 'solicitudes', f'Solicitud editada: id={id}', id)
            flash('Datos actualizados.','success')
            return redirect(url_for('solicitudes.ver', id=id))
        except Exception as e:
            conn.rollback(); flash(f'Error: {str(e)}','danger')
        finally:
            cur.close(); conn.close()

    return render_template('solicitudes/formulario.html', s=s, editando=True)


# ─────────────────────────────────────────────────────────────────
# CONVALIDACIÓN DE CURSOS
# ─────────────────────────────────────────────────────────────────

@bp.route('/convalidar/<int:id>')
def convalidar(id):
    try:
        s = get_solicitud_completa(id)
        if not s:
            flash('Solicitud no encontrada.', 'danger')
            return redirect(url_for('solicitudes.index'))
        return render_template('solicitudes/convalidacion.html', s=s)
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f'Error al cargar convalidación: {e}', 'danger')
        return redirect(url_for('solicitudes.index'))


@bp.route('/guardar-convalidacion/<int:id>', methods=['POST'])
def guardar_convalidacion(id):
    data = request.get_json()
    conn = None; cur = None
    try:
        conn = get_connection(); cur = conn.cursor()
        # Actualizar planes seleccionados
        cur.execute("""
            UPDATE solicitudes SET plan_local_id=%s, plan_externo_id=%s WHERE id=%s
        """, (data.get('plan_local_id'), data.get('plan_externo_id'), id))

        # Reemplazar cursos
        cur.execute("DELETE FROM solicitud_cursos WHERE solicitud_id=%s", (id,))
        for item in data.get('cursos', []):
            cur.execute("""
                INSERT INTO solicitud_cursos
                    (solicitud_id, curso_local_id, curso_externo_id, nota, estado, periodo_lectivo)
                VALUES (%s,%s,%s,%s,%s,%s)
            """, (id, item['curso_local_id'], item.get('curso_externo_id'),
                  item.get('nota'), item.get('estado','pendiente'), item.get('periodo_lectivo')))

        conn.commit()
        invalidar_cache(id)
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()


@bp.route('/curso/eliminar/<int:curso_id>', methods=['POST'])
def eliminar_curso(curso_id):
    conn = None; cur = None
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT solicitud_id FROM solicitud_cursos WHERE id=%s", (curso_id,))
        row = cur.fetchone()
        cur.execute("DELETE FROM solicitud_cursos WHERE id=%s", (curso_id,))
        conn.commit()
        if row: invalidar_cache(row[0])
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()


@bp.route('/curso/editar/<int:curso_id>', methods=['POST'])
def editar_curso(curso_id):
    data = request.get_json()
    conn = None; cur = None
    try:
        conn = get_connection(); cur = conn.cursor()
        cur.execute("SELECT solicitud_id FROM solicitud_cursos WHERE id=%s", (curso_id,))
        row = cur.fetchone()
        cur.execute("""
            UPDATE solicitud_cursos
            SET nota=%s, periodo_lectivo=%s
            WHERE id=%s
        """, (data.get('nota'), data.get('periodo_lectivo'), curso_id))
        conn.commit()
        if row: invalidar_cache(row[0])
        return jsonify({'ok': True})
    except Exception as e:
        if conn: conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        if cur: cur.close()
        if conn: conn.close()


# ─────────────────────────────────────────────────────────────────
# APIs
# ─────────────────────────────────────────────────────────────────

@bp.route('/api/buscar-postulante')
def api_buscar_postulante():
    q = request.args.get('q','').strip()
    if len(q) < 2: return jsonify([])
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    like = f'%{q}%'
    cur.execute("""
        SELECT id, codigo, apellidos_nombres, dni, programa, sexo,
               modalidad_estudios, modalidad_admision, semestre_academico,
               turno, asesora, correo, celular, institucion_procedencia
        FROM postulantes
        WHERE apellidos_nombres ILIKE %s OR dni ILIKE %s OR codigo ILIKE %s OR programa ILIKE %s
        LIMIT 10
    """, (like, like, like,like))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify(rows)


@bp.route('/api/planes-por-tipo')
def api_planes_por_tipo():
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT p.id, p.nombre_plan, p.tipo_plan, p.periodo_academico,
               COUNT(c.id) AS total_cursos, COALESCE(SUM(c.creditos),0) AS total_creditos
        FROM planes_estudio p
        LEFT JOIN cursos_plan c ON c.plan_id = p.id
        GROUP BY p.id ORDER BY p.nombre_plan, p.periodo_academico
    """)
    planes = cur.fetchall(); cur.close(); conn.close()
    grupos = {}
    for p in planes:
        n = p['nombre_plan']
        if n not in grupos:
            grupos[n] = {'nombre': n, 'tipo': p['tipo_plan'], 'periodos': []}
        grupos[n]['periodos'].append({
            'id': p['id'], 'periodo': p['periodo_academico'],
            'total_cursos': p['total_cursos'], 'total_creditos': p['total_creditos']
        })
    locales  = [g for g in grupos.values() if g['tipo'] == 'local']
    externos = [g for g in grupos.values() if g['tipo'] == 'externo']
    return jsonify({'locales': locales, 'externos': externos})


@bp.route('/api/cursos-plan/<int:plan_id>')
def api_cursos_plan(plan_id):
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    cur.execute(f"""
        SELECT id, ciclo, codigo, nombre_curso, creditos, prerrequisito
        FROM cursos_plan WHERE plan_id=%s
        ORDER BY {_ORDEN_CICLO}, nombre_curso
    """, (plan_id,))
    cursos = cur.fetchall(); cur.close(); conn.close()
    return jsonify(cursos)

@bp.route('/eliminar/<int:id>', methods=['POST'])
def eliminar(id):
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM solicitudes WHERE id=%s", (id,))
        conn.commit()
        invalidar_cache(id)
        registrar('eliminar', 'solicitudes', f'Solicitud eliminada: id={id}', id)
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback(); return jsonify({'ok': False, 'error': str(e)}), 500
    finally: cur.close(); conn.close()


@bp.route('/api/verificar-duplicado')
def verificar_duplicado():
    postulante_id = request.args.get('postulante_id', type=int)
    excluir_id    = request.args.get('excluir_id', type=int)
    if not postulante_id:
        return jsonify({'existe': False})
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    sql = "SELECT s.id, s.codigo FROM solicitudes s WHERE s.postulante_id=%s"
    params = [postulante_id]
    if excluir_id:
        sql += " AND s.id != %s"
        params.append(excluir_id)
    cur.execute(sql, params)
    row = cur.fetchone()
    cur.close(); conn.close()
    return jsonify({'existe': bool(row), 'solicitud': row})

@bp.route('/api/cursos-guardados/<int:id>')
def api_cursos_guardados(id):
    """Devuelve los cursos guardados de una solicitud para precargar la convalidación."""
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT sc.curso_local_id, sc.curso_externo_id, sc.nota, sc.estado, sc.periodo_lectivo
        FROM solicitud_cursos sc
        WHERE sc.solicitud_id = %s
    """, (id,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    return jsonify(rows)

@bp.route('/marcar-emitido/<int:id>', methods=['POST'])
def marcar_emitido(id):
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("UPDATE solicitudes SET estado='emitido' WHERE id=%s", (id,))
        conn.commit()
        registrar('editar', 'solicitudes', f'Solicitud marcada como emitida: id={id}', id)
        flash('Solicitud marcada como emitida.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {str(e)}','danger')
    finally:
        cur.close(); conn.close()
    return redirect(url_for('solicitudes.ver', id=id))


@bp.route('/enviar-correo/<int:id>', methods=['POST'])
def enviar_correo(id):
    from routes.correos import enviar_correo as enviar_email
    from routes.generar_word import generar_pdf
    from datetime import datetime
    
    data = request.get_json() or {}
    destinatario = data.get('correo', '').strip()
    asunto = data.get('asunto', f'Resolucion de Convalidacion - Solicitud {id}')
    mensaje = data.get('mensaje', '')
    adjuntar_pdf = data.get('adjuntar_pdf', True)
    
    if not destinatario or '@' not in destinatario:
        return jsonify({'ok': False, 'error': 'Correo invalido'})
    
    conn = None; cur = None
    try:
        conn = get_connection(); cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT s.*, COALESCE(p.apellidos_nombres,'') AS nombre,
                   COALESCE(p.correo,'') AS correo_postulant,
                   COALESCE(p.programa,'') AS programa
            FROM solicitudes s
            LEFT JOIN postulantes p ON s.postulante_id = p.id
            WHERE s.id=%s
        """, (id,))
        sol = cur.fetchone()
        
        if not sol:
            return jsonify({'ok': False, 'error': 'Solicitud no encontrada'})
        
        if sol['estado'] != 'emitido':
            return jsonify({'ok': False, 'error': 'Solo se pueden enviar solicitudes emitidas'})
        
        buffer_pdf, nombre_adjunto = generar_pdf(id)
        buffer_pdf.seek(0)
        ruta_adjunto_suffix = '.pdf'
        
        import tempfile, os
        with tempfile.NamedTemporaryFile(suffix=ruta_adjunto_suffix, delete=False) as tmp:
            tmp.write(buffer_pdf.read())
            ruta_adjunto = tmp.name
        
        token = secrets.token_urlsafe(32)
        base_url = request.host_url.rstrip('/')
        
        # URL para confirmacion - usar variable de entorno o detectar automaticamente
        confirm_base_url = os.getenv('APP_BASE_URL', base_url)
        
        url_confirmar = f"{confirm_base_url}/solicitudes/confirmar/{token}"
        url_rechazar = f"{confirm_base_url}/solicitudes/rechazar/{token}"
        
        cur.execute("""
            UPDATE solicitudes 
            SET token_confirmacion = %s, 
                estado_confirmacion = 'pendiente'
            WHERE id = %s
        """, (token, id))
        conn.commit()
        
        template_id = data.get('template_id')
        if template_id:
            cur.execute("SELECT cuerpo FROM plantillas_correo WHERE id=%s AND activo", (template_id,))
            row = cur.fetchone()
            cuerpo_base = row['cuerpo'] if row else ''
        else:
            # Si no hay plantilla, usar el mensaje digitado tal cual (texto plano o HTML)
            cuerpo_base = mensaje
        
        def render_html(template, **vars):
            result = template
            for k, v in vars.items():
                result = result.replace('@' + k, str(v))
                result = result.replace('{{' + k + '}}', str(v))
            return result
        
        cuerpo_plantilla = render_html(
            cuerpo_base,
            codigo=sol['codigo'],
            nombre=sol['nombre'],
            programa=sol['programa'] or '—'
        )
        
        botones_confirmacion = f'''
        <div style="background: #fffbeb; border: 2px solid #f59e0b; border-radius: 8px; padding: 20px; margin-top: 30px; text-align: center;">
            <p style="font-weight: 600; margin-bottom: 15px; color: #92400e;">¿Esta de acuerdo con esta resolucion?</p>
            <p style="font-size: 13px; color: #666; margin-bottom: 20px;">Haga clic en el boton correspondiente para confirmar o rechazar</p>
            <div style="display: flex; gap: 15px; justify-content: center; flex-wrap: wrap;">
                <a href="{url_confirmar}" style="background: #16a34a; color: white; padding: 12px 25px; border-radius: 6px; text-decoration: none; font-weight: 600; font-size: 14px;">✔ Confirmar</a>
                <a href="{url_rechazar}" style="background: #dc2626; color: white; padding: 12px 25px; border-radius: 6px; text-decoration: none; font-weight: 600; font-size: 14px;">✕ Rechazar</a>
            </div>
        </div>'''
        
        cuerpo_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333; max-width: 600px; margin: 0 auto;">
            <div style="background: linear-gradient(135deg, #1F3864, #4a7cc7); padding: 20px; text-align: center;">
                <h1 style="color: white; margin: 0; font-size: 24px;">Resolucion de Convalidacion</h1>
            </div>
            <div style="padding: 30px;">
                {cuerpo_plantilla}
                {botones_confirmacion}
            </div>
        </body>
        </html>
        """
        
        adjuntos = [(ruta_adjunto, nombre_adjunto)] if adjuntar_pdf else None
        
        resultado = enviar_email(
            destinatario=destinatario,
            asunto=asunto,
            cuerpo_html=cuerpo_html,
            adjuntos=adjuntos,
            usuario_id=session.get('usuario_id')
        )
        
        try:
            if os.path.exists(ruta_adjunto):
                os.remove(ruta_adjunto)
        except:
            pass
        
        if resultado['ok']:
            registrar('enviar_correo', 'solicitudes', f'Correo enviado a {destinatario} para solicitud {sol["codigo"]} con confirmacion', id)
            return jsonify({'ok': True})
        else:
            return jsonify({'ok': False, 'error': resultado.get('error', 'Error desconocido')})
    
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        if cur: cur.close()
        if conn: conn.close()


@bp.route('/correo-preview/<int:id>')
def correo_preview(id):
    """Devuelve los datos para el modal de envio de correo"""
    from routes.correos import get_config_correo, get_plantillas, renderizar_plantilla
    from routes.generar_word import generar_pdf
    from datetime import datetime
    
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.*, COALESCE(p.apellidos_nombres,'') AS nombre,
                   COALESCE(p.correo,'') AS correo_postulant,
                   COALESCE(p.celular,'') AS celular,
                   COALESCE(p.dni,'') AS dni,
                   COALESCE(p.programa,'') AS programa,
                   COALESCE(p.modalidad_estudios,'') AS modalidad,
                   COALESCE(NULLIF(p.institucion_procedencia,''), pe.nombre_plan, '') AS ies_origen
            FROM solicitudes s
            LEFT JOIN postulantes p ON s.postulante_id = p.id
            LEFT JOIN planes_estudio pe ON s.plan_externo_id = pe.id
            WHERE s.id=%s
        """, (id,))
        sol = cur.fetchone()
        
        if not sol:
            return jsonify({'ok': False, 'error': 'Solicitud no encontrada'})
        
        if sol['estado'] != 'emitido':
            return jsonify({'ok': False, 'error': 'Solo se pueden enviar solicitudes emitidas'})
        
        es_admin = session.get('usuario_rol') == 'admin'
        usuario_id = session.get('usuario_id')
        
        if es_admin:
            cur.execute("""
                SELECT c.id, c.correo_remitente, c.nombre_remitente, c.activo, u.nombre_completo as usuario_nombre
                FROM config_correo c
                LEFT JOIN usuarios u ON c.usuario_id = u.id
                WHERE c.correo_remitente!=''
                ORDER BY c.activo DESC, c.fecha_creacion DESC
            """)
            configs = cur.fetchall()
        else:
            cur.execute("""
                SELECT id, correo_remitente, nombre_remitente, activo, NULL as usuario_nombre
                FROM config_correo 
                WHERE usuario_id=%s AND correo_remitente!=''
                ORDER BY activo DESC, fecha_creacion DESC
            """, (usuario_id,))
            configs = cur.fetchall()
            
            if not configs:
                cur.execute("""
                    SELECT id, correo_remitente, nombre_remitente, activo, 'Admin' as usuario_nombre
                    FROM config_correo 
                    WHERE activo AND correo_remitente!='' AND usuario_id=1
                    LIMIT 1
                """)
                configs = cur.fetchall()
        
        plantillas = get_plantillas()
        
        buffer_pdf, nombre_pdf = generar_pdf(id)
        buffer_pdf.seek(0)
        pdf_base64 = base64.b64encode(buffer_pdf.read()).decode('utf-8')
        
        fecha = sol['fecha_emision'].strftime('%d/%m/%Y') if sol.get('fecha_emision') else datetime.now().strftime('%d/%m/%Y')
        
        datos = {
            'codigo': sol['codigo'],
            'nombre': sol['nombre'],
            'dni': sol.get('dni', ''),
            'programa': sol.get('programa', ''),
            'modalidad': sol.get('modalidad', ''),
            'ies_origen': sol.get('ies_origen', ''),
            'fecha': fecha,
            'total_costo': str(sol.get('total_costo', '')),
            'correo': sol['correo_postulant'],
            'celular': sol.get('celular', '')
        }
        
        default_html_template = '''<p>Hola <strong>@nombre</strong></p><br><br><p>Bienvenido(a) al Sistema de Convalidaciones UAI.</p>
<div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 15px; margin: 20px 0;">
    <p style="margin: 5px 0;"><strong>Codigo de solicitud:</strong> @codigo</p>
    <p style="margin: 5px 0;"><strong>Programa:</strong> @programa</p>
</div>
<p>Su reporte de convalidacion ha sido generado y se encuentra disponible en este correo.</p>
<p><strong>Por favor, revise el documento adjunto.</strong></p>
<p style="margin-top: 25px; color: #666; font-size: 12px;">
    Para cualquier consulta, comunicarse con la oficina de Convalidaciones.<br>
    Saludos cordiales,<br>
    <strong>Sistema de Convalidaciones UAI</strong>
</p>'''
        
        return jsonify({
            'ok': True,
            'configs': configs,
            'plantillas': [{'id': p['id'], 'nombre': p['nombre'], 'asunto': p['asunto'], 'cuerpo': p['cuerpo']} for p in plantillas],
            'default_html_template': default_html_template,
            'postulante': {
                'correo': sol['correo_postulant'],
                'nombre': sol['nombre'],
                'codigo': sol['codigo'],
                'dni': sol.get('dni', ''),
                'programa': sol.get('programa', ''),
                'modalidad': sol.get('modalidad', '')
            },
            'datos': datos,
            'pdf_nombre': nombre_pdf,
            'pdf_preview': pdf_base64
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        cur.close()
        conn.close()


@bp.route('/whatsapp-preview/<int:id>')
def whatsapp_preview(id):
    """Devuelve datos para el modal de envío de WhatsApp"""
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.*, COALESCE(p.apellidos_nombres,'') AS nombre,
                   COALESCE(p.correo,'') AS correo_postulant,
                   COALESCE(p.celular,'') AS celular,
                   COALESCE(p.programa,'') AS programa
            FROM solicitudes s
            LEFT JOIN postulantes p ON s.postulante_id = p.id
            WHERE s.id=%s
        """, (id,))
        sol = cur.fetchone()
        
        if not sol:
            return jsonify({'ok': False, 'error': 'Solicitud no encontrada'})
        
        if sol['estado'] != 'emitido':
            return jsonify({'ok': False, 'error': 'Solo se pueden enviar solicitudes emitidas'})
        
        cur.execute("""
            SELECT id FROM logs_sistema 
            WHERE modulo='solicitudes' AND entidad_id=%s AND accion='enviar_correo'
            LIMIT 1
        """, (id,))
        correo_enviado = cur.fetchone() is not None
        
        genero = ''
        nombre_parts = sol['nombre'].split(' ')
        if len(nombre_parts) >= 2:
            primer_nombre = nombre_parts[0].lower()
            if primer_nombre.endswith('a') and not primer_nombre.endswith('ea'):
                genero = 'a'
            elif primer_nombre.endswith('o') or primer_nombre.endswith('e'):
                genero = 'o'
        
        correo_mensaje = f"y enviada a su correo {sol['correo_postulant']}" if sol['correo_postulant'] else ""
        
        celular_raw = sol['celular'] or ''
        celular_limpio = ''.join(c for c in celular_raw if c.isdigit())
        if celular_limpio.startswith('9') and len(celular_limpio) == 9:
            celular_formato = '51' + celular_limpio
        elif celular_limpio.startswith('51') and len(celular_limpio) >= 11:
            celular_formato = celular_limpio
        else:
            celular_formato = '51' + celular_limpio if len(celular_limpio) >= 9 else ''
        
        mensaje = f"""Hola {sol['nombre']},\n\nBienvenido(a) al Sistema de Convalidaciones UAI.\n\n📋 Codigo de solicitud: {sol['codigo']}\n🎓 Programa: {sol['programa']}\n\nSu reporte de convalidacion ha sido enviado a su correo electronico registrado: {sol['correo_postulant']}\n\n📧 Revise su bandeja de entrada (incluyendo spam).\n\nPara cualquier consulta, concurrentese con la oficina de Convalidaciones.\n\nSaludos cordiales,\nSistema de Convalidaciones UAI"""
        
        return jsonify({
            'ok': True,
            'postulante': {
                'nombre': sol['nombre'],
                'celular': celular_formato,
                'correo': sol['correo_postulant'],
                'codigo': sol['codigo'],
                'programa': sol['programa'],
                'dni': sol.get('dni', '')
            },
            'correo_enviado': correo_enviado,
            'mensaje': mensaje
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        cur.close()
        conn.close()


@bp.route('/wpp-log/<int:id>', methods=['POST'])
def wpp_log(id):
    """Registra el envío de WhatsApp en logs"""
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT codigo FROM solicitudes WHERE id=%s", (id,))
        sol = cur.fetchone()
        if sol:
            registrar('enviar_whatsapp', 'solicitudes', f'WhatsApp enviado para solicitud {sol["codigo"]}', id)
        return jsonify({'ok': True})
    except:
        return jsonify({'ok': False})
    finally:
        cur.close()
        conn.close()


# ===== RUTAS DE CONFIRMACION =====
@bp.route('/confirmar/<token>')
def confirmar_solicitud(token):
    """Procesa la confirmacion de una solicitud via token"""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.*, p.apellidos_nombres AS nombre, p.correo AS correo_postulant,
                   p.dni, p.programa
            FROM solicitudes s
            LEFT JOIN postulantes p ON s.postulante_id = p.id
            WHERE s.token_confirmacion = %s
        """, (token,))
        sol = cur.fetchone()
        
        if not sol:
            return render_template('solicitudes/confirmacion.html', 
                                 estado='error', 
                                 mensaje='Enlace de confirmacion invalido o expirado.')
        
        if sol['confirmado'] == 1:
            return render_template('solicitudes/confirmacion.html',
                                 estado='info',
                                 mensaje='Esta solicitud ya fue confirmada anteriormente.')
        
        if sol['confirmado'] == 2:
            return render_template('solicitudes/confirmacion.html',
                                 estado='info',
                                 mensaje='Esta solicitud fue rechazada anteriormente.')
        
        cur.execute("""
            UPDATE solicitudes 
            SET confirmado = 1,
                estado_confirmacion = 'aceptado',
                fecha_confirmacion = NOW()
            WHERE id = %s
        """, (sol['id'],))
        conn.commit()
        
        registrar('confirmacion', 'solicitudes', 
                 f'Postulante confirmo la solicitud {sol["codigo"]}', sol['id'])
        
        return render_template('solicitudes/confirmacion.html',
                             estado='success',
                             mensaje='Su confirmacion ha sido registrada correctamente.',
                             datos=sol)
    except Exception as e:
        return render_template('solicitudes/confirmacion.html',
                             estado='error',
                             mensaje=f'Error al procesar la confirmacion: {str(e)}')
    finally:
        cur.close()
        conn.close()


@bp.route('/rechazar/<token>', methods=['GET', 'POST'])
def rechazar_solicitud(token):
    """Procesa el rechazo de una solicitud via token"""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    
    if request.method == 'POST':
        motivo = request.form.get('motivo', '').strip()
        cur.execute("""
            SELECT s.*, p.apellidos_nombres AS nombre, p.correo AS correo_postulant,
                   p.dni, p.programa
            FROM solicitudes s
            LEFT JOIN postulantes p ON s.postulante_id = p.id
            WHERE s.token_confirmacion = %s
        """, (token,))
        sol = cur.fetchone()
        
        if not sol:
            return render_template('solicitudes/confirmacion.html',
                                 estado='error',
                                 mensaje='Enlace de confirmacion invalido o expirado.')
        
        if sol['confirmado'] in (1, 2):
            return render_template('solicitudes/confirmacion.html',
                                 estado='info',
                                 mensaje='Esta solicitud ya fue procesada anteriormente.')
        
        cur.execute("""
            UPDATE solicitudes 
            SET confirmado = 2,
                estado_confirmacion = 'rechazado',
                fecha_confirmacion = NOW(),
                observacion = CONCAT(COALESCE(observacion, ''), ' | Rechazo: ', %s)
            WHERE id = %s
        """, (motivo, sol['id']))
        conn.commit()
        
        registrar('rechazo', 'solicitudes',
                 f'Postulante rechazo la solicitud {sol["codigo"]}. Motivo: {motivo}', sol['id'])
        
        return render_template('solicitudes/confirmacion.html',
                             estado='rechazado',
                             mensaje='Su rechazo ha sido registrado. Nos pondremos en contacto con usted.',
                             datos=sol)
    
    cur.execute("""
        SELECT s.*, p.apellidos_nombres AS nombre, p.correo AS correo_postulant,
               p.dni, p.programa
        FROM solicitudes s
        LEFT JOIN postulantes p ON s.postulante_id = p.id
        WHERE s.token_confirmacion = %s
    """, (token,))
    sol = cur.fetchone()
    
    if not sol:
        return render_template('solicitudes/confirmacion.html',
                             estado='error',
                             mensaje='Enlace de confirmacion invalido o expirado.')
    
    if sol['confirmado'] == 1:
        return render_template('solicitudes/confirmacion.html',
                             estado='info',
                             mensaje='Esta solicitud ya fue confirmada anteriormente.')
    
    if sol['confirmado'] == 2:
        return render_template('solicitudes/confirmacion.html',
                             estado='info',
                             mensaje='Esta solicitud fue rechazada anteriormente.')
    
    cur.close()
    conn.close()
    return render_template('solicitudes/rechazo.html', sol=sol, token=token)


@bp.route('/confirmacion-data/<int:id>')
def confirmacion_data(id):
    """Devuelve los datos de confirmacion de una solicitud"""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.id, s.codigo, s.confirmado, s.fecha_confirmacion, s.observacion
            FROM solicitudes s
            WHERE s.id = %s
        """, (id,))
        sol = cur.fetchone()
        
        if not sol:
            return jsonify({'ok': False, 'error': 'Solicitud no encontrada'})
        
        fecha_str = ''
        if sol.get('fecha_confirmacion'):
            fecha_str = sol['fecha_confirmacion'].strftime('%d/%m/%Y %H:%M')
        
        return jsonify({
            'ok': True,
            'codigo': sol['codigo'],
            'confirmado': sol['confirmado'],
            'fecha_confirmacion': fecha_str,
            'observacion': sol.get('observacion') or ''
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        cur.close()
        conn.close()


@bp.route('/consolidado-excel/<int:id>')
def consolidado_excel(id):
    """Genera y descarga un Excel con el consolidado de convalidacion."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.*,
                   COALESCE(p.apellidos_nombres, '') AS nombre,
                   COALESCE(p.dni, '') AS dni,
                   COALESCE(p.programa, '') AS programa,
                   COALESCE(p.institucion_procedencia, '') AS institucion,
                   COALESCE(p.modalidad_admision, '') AS modalidad,
                   COALESCE(c.nombre, '') AS carrera_nombre
              FROM solicitudes s
              LEFT JOIN postulantes p ON s.postulante_id = p.id
              LEFT JOIN carreras c ON s.carrera_id = c.id
             WHERE s.id = %s
        """, (id,))
        sol = cur.fetchone()
        if not sol:
            return 'Solicitud no encontrada', 404

        _oc = "CASE cp_e.ciclo WHEN 'I' THEN 1 WHEN 'II' THEN 2 WHEN 'III' THEN 3 WHEN 'IV' THEN 4 WHEN 'V' THEN 5 WHEN 'VI' THEN 6 WHEN 'VII' THEN 7 WHEN 'VIII' THEN 8 WHEN 'IX' THEN 9 WHEN 'X' THEN 10 END"
        cur.execute(f"""
            SELECT cp_e.id AS curso_id, cp_e.ciclo, cp_e.codigo AS ext_codigo,
                   cp_e.nombre_curso AS ext_nombre, cp_e.creditos AS ext_creditos,
                   COALESCE(cp_e.prerrequisito, '') AS prerrequisito,
                   COALESCE(cp_l.nombre_curso, '') AS local_nombre,
                   cp_l.creditos AS local_creditos,
                   sc.nota, COALESCE(sc.estado, 'sin_validar') AS estado,
                   COALESCE(sc.periodo_lectivo, '') AS periodo_lectivo
              FROM cursos_plan cp_e
              LEFT JOIN solicitud_cursos sc ON sc.curso_externo_id = cp_e.id AND sc.solicitud_id = %s
              LEFT JOIN cursos_plan cp_l ON sc.curso_local_id = cp_l.id
             WHERE cp_e.plan_id = %s
             ORDER BY {_oc}, cp_e.nombre_curso
        """, (id, sol.get('plan_externo_id')))
        cursos = cur.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Consolidado'

        title_font = Font(bold=True, size=14, color='1F3864')
        subtitle_font = Font(bold=True, size=11, color='1F3864')
        header_font = Font(bold=True, size=10, color='FFFFFF')
        normal_font = Font(size=10)
        bold_font = Font(bold=True, size=10)
        total_font = Font(bold=True, size=11, color='1F3864')

        header_fill = PatternFill('solid', fgColor='1F3864')
        section_fill = PatternFill('solid', fgColor='D9E2F0')
        light_gray_fill = PatternFill('solid', fgColor='F2F2F2')

        thin_border = Border(
            left=Side(style='thin', color='999999'),
            right=Side(style='thin', color='999999'),
            top=Side(style='thin', color='999999'),
            bottom=Side(style='thin', color='999999'),
        )
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Colores por periodo
        period_palette = [
            'D6EAF8', 'D5F5E3', 'FDEBD0', 'E8DAEF', 'FADBD8',
            'D1F2EB', 'FCF3CF', 'F5CBA7', 'AED6F1', 'A9DFBF'
        ]
        distinct_periodos = {}
        period_idx = 0
        for c in cursos:
            p = c.get('periodo_lectivo', '') or ''
            if p and p not in distinct_periodos:
                distinct_periodos[p] = period_palette[period_idx % len(period_palette)]
                period_idx += 1

        # ── TITULO ──
        ws.merge_cells('A1:J1')
        ws['A1'] = 'REPORTE DE CONVALIDACIÓN'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # ── CABECERA ──
        row = 3
        for lbl, val in [('Estudiante:', sol.get('nombre', '')), ('Modalidad:', sol.get('modalidad', ''))]:
            ws.cell(row=row, column=1, value=lbl).font = bold_font
            ws.cell(row=row, column=2, value=val).font = normal_font
            row += 1

        ws.cell(row=row, column=1, value='Código:').font = bold_font
        ws.cell(row=row, column=2, value=sol.get('codigo', '')).font = normal_font
        row += 1
        ws.cell(row=row, column=1, value='IES:').font = bold_font
        ws.cell(row=row, column=2, value=sol.get('institucion', '')).font = normal_font
        row += 1
        ws.cell(row=row, column=1, value='PLAN DE ESTUDIOS:').font = bold_font
        ws.cell(row=row, column=2, value=sol.get('carrera_nombre', '')).font = subtitle_font
        row += 2

        # ── ENCABEZADOS DE TABLA ──
        # Secciones: Ext (1-5) | Convalidables (6-8) | Estado (9) | Periodo (10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value='ASIGNATURAS DEL PLAN EXTERNO').font = Font(bold=True, size=10, color='1F3864')
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).alignment = center_align
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = section_fill

        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        ws.cell(row=row, column=6, value='CONVALIDACIÓN').font = Font(bold=True, size=10, color='1F3864')
        ws.cell(row=row, column=6).fill = section_fill
        ws.cell(row=row, column=6).alignment = center_align
        for c in range(6, 9):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = section_fill

        row += 1

        # Sub-encabezados
        sub_headers = ['CICLO', 'CÓDIGO', 'NOMBRE DEL CURSO', 'CRÉDITOS', 'PRERREQUISITO',
                       'NOMBRE DEL CURSO', 'CRÉDITOS', 'NOTA']
        for i, h in enumerate(sub_headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border

        ws.row_dimensions[row].height = 30
        row += 1

        # ── DATOS ──
        data_start = row
        total_creditos = 0
        for idx, c in enumerate(cursos):
            ext_cred = c.get('ext_creditos', 0) or 0
            total_creditos += ext_cred
            estado = c.get('estado', 'sin_validar')

            # Period fill
            periodo = c.get('periodo_lectivo', '') or ''
            period_fill = None
            if periodo and periodo in distinct_periodos:
                period_fill = PatternFill('solid', fgColor=distinct_periodos[periodo])

            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

            ws.cell(row=row, column=1, value=c.get('ciclo', '')).font = normal_font
            ws.cell(row=row, column=1).alignment = center_align
            ws.cell(row=row, column=2, value=c.get('ext_codigo', '')).font = normal_font
            ws.cell(row=row, column=2).alignment = center_align
            ws.cell(row=row, column=3, value=c.get('ext_nombre', '')).font = normal_font
            ws.cell(row=row, column=3).alignment = left_align
            ws.cell(row=row, column=4, value=ext_cred).font = normal_font
            ws.cell(row=row, column=4).alignment = center_align
            ws.cell(row=row, column=5, value=c.get('prerrequisito', '')).font = normal_font
            ws.cell(row=row, column=5).alignment = center_align

            # Nombre curso convalidacion segun estado
            if estado == 'convalidado':
                curso_conv = c.get('local_nombre', '')
            elif estado == 'examen_suficiencia':
                curso_conv = 'EXAMEN SUFICIENCIA'
            elif estado == 'pendiente':
                curso_conv = periodo if periodo else '—'
            else:
                curso_conv = periodo if periodo else '—'
            ws.cell(row=row, column=6, value=curso_conv).font = normal_font
            ws.cell(row=row, column=6).alignment = left_align
            ws.cell(row=row, column=7, value=c.get('local_creditos', 0) or 0).font = normal_font
            ws.cell(row=row, column=7).alignment = center_align

            # Nota: solo para convalidado
            if estado == 'convalidado':
                nota = c.get('nota')
                nota_val = float(nota) if nota is not None else None
                nota_str = str(int(nota_val)) if nota_val is not None and nota_val == int(nota_val) else (str(nota_val) if nota_val is not None else '-')
            else:
                nota_str = '-'
            nota_cell = ws.cell(row=row, column=8, value=nota_str)
            nota_cell.font = normal_font
            nota_cell.alignment = center_align

            # Aplicar color de periodo si existe
            if period_fill:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = period_fill
            elif idx % 2 == 1:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = light_gray_fill

            row += 1

        # ── TOTAL ──
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row=row, column=1, value='TOTAL DE CRÉDITOS DEL PROGRAMA:')
        ws.cell(row=row, column=1).font = total_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=1).border = thin_border
        for c in range(2, 9):
            ws.cell(row=row, column=c).border = thin_border

        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws.cell(row=row, column=6, value=total_creditos)
        ws.cell(row=row, column=6).font = total_font
        ws.cell(row=row, column=6).alignment = center_align
        ws.cell(row=row, column=8).border = thin_border

        # ── ANCHO DE COLUMNAS ──
        col_widths = [8, 12, 38, 9, 16, 38, 9, 8]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # ── CONGELAR ENCABEZADOS ──
        ws.freeze_panes = f'A{data_start}'

        # ── CONFIGURACIÓN DE IMPRESIÓN ──
        ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.orientation = 'landscape'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        nombre_archivo = f"Reporte_Convalidacion_{sol.get('codigo', 'sin_codigo')}.xlsx"

        return send_file(
            buf,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return f'Error: {str(e)}', 500
    finally:
        cur.close()
        conn.close()


@bp.route('/consolidado-preview/<int:id>')
def consolidado_preview(id):
    """Genera el consolidado de convalidacion en HTML (para preview en iframe)"""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.*,
                   COALESCE(p.apellidos_nombres, '') AS nombre,
                   COALESCE(p.dni, '') AS dni,
                   COALESCE(p.programa, '') AS programa,
                   COALESCE(p.institucion_procedencia, '') AS institucion,
                   COALESCE(p.modalidad_admision, '') AS modalidad,
                   COALESCE(c.nombre, '') AS carrera_nombre,
                   pe.nombre_plan AS plan_nombre
              FROM solicitudes s
              LEFT JOIN postulantes p ON s.postulante_id = p.id
              LEFT JOIN carreras c ON s.carrera_id = c.id
              LEFT JOIN planes_estudio pe ON s.plan_externo_id = pe.id
             WHERE s.id = %s
        """, (id,))
        sol = cur.fetchone()
        if not sol:
            return '<div style="padding:2rem;text-align:center;color:#dc2626;">Solicitud no encontrada</div>', 404

        _oc = "CASE cp_e.ciclo WHEN 'I' THEN 1 WHEN 'II' THEN 2 WHEN 'III' THEN 3 WHEN 'IV' THEN 4 WHEN 'V' THEN 5 WHEN 'VI' THEN 6 WHEN 'VII' THEN 7 WHEN 'VIII' THEN 8 WHEN 'IX' THEN 9 WHEN 'X' THEN 10 END"
        cur.execute(f"""
            SELECT cp_e.ciclo, cp_e.codigo AS ext_codigo,
                   cp_e.nombre_curso AS ext_nombre, cp_e.creditos AS ext_creditos,
                   COALESCE(cp_e.prerrequisito, '') AS prerrequisito,
                   COALESCE(cp_l.nombre_curso, '') AS local_nombre,
                   cp_l.creditos AS local_creditos,
                   sc.nota, COALESCE(sc.estado, 'sin_validar') AS estado,
                   COALESCE(sc.periodo_lectivo, '') AS periodo_lectivo
              FROM cursos_plan cp_e
              LEFT JOIN solicitud_cursos sc ON sc.curso_externo_id = cp_e.id AND sc.solicitud_id = %s
              LEFT JOIN cursos_plan cp_l ON sc.curso_local_id = cp_l.id
             WHERE cp_e.plan_id = %s
             ORDER BY {_oc}, cp_e.nombre_curso
        """, (id, sol.get('plan_externo_id')))
        cursos = cur.fetchall()

        total_creditos = sum(c.get('ext_creditos', 0) or 0 for c in cursos)

        # Colores por periodo
        period_palette = [
            'D6EAF8', 'D5F5E3', 'FDEBD0', 'E8DAEF', 'FADBD8',
            'D1F2EB', 'FCF3CF', 'F5CBA7', 'AED6F1', 'A9DFBF'
        ]
        distinct_periodos = {}
        period_idx = 0
        for c in cursos:
            p = c.get('periodo_lectivo', '') or ''
            if p and p not in distinct_periodos:
                distinct_periodos[p] = period_palette[period_idx % len(period_palette)]
                period_idx += 1

        period_css = ''
        for p, bg in distinct_periodos.items():
            period_css += f'.period-{p.replace("-","").replace("/","")} {{ background:#{bg}; }}\n'

        html = f"""
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ font-family:Arial,sans-serif; font-size:9px; background:#fff; }}
.container {{ width:210mm; min-height:297mm; margin:0 auto; padding:8mm 7mm; background:white; }}
.header {{ text-align:center; margin-bottom:6px; border-bottom:2px solid #1F3864; padding-bottom:5px; }}
.header h1 {{ color:#1F3864; font-size:13px; margin-bottom:2px; }}
.header h2 {{ color:#4a7cc7; font-size:9px; font-weight:normal; }}
.info-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:3px 20px; margin-bottom:8px; font-size:8.5px; }}
.info-item {{}}
.info-label {{ font-weight:bold; color:#555; }}
table {{ width:100%; border-collapse:collapse; margin-bottom:6px; font-size:7.5px; }}
th {{ background:#1F3864; color:white; padding:3px 2px; text-align:center; font-weight:600; font-size:7px; }}
td {{ padding:2px 3px; border:1px solid #ccc; }}
.num {{ text-align:center; }}
{period_css}
.footer {{ text-align:center; font-size:6.5px; color:#888; margin-top:8px; padding-top:4px; border-top:1px solid #eee; }}
@page {{ size:A4; margin:0; }}
</style>
<div class="container">
  <div class="header">
    <h1>REPORTE DE CONVALIDACIÓN</h1>
    <h2>{sol.get('carrera_nombre', '')}</h2>
  </div>

  <div class="info-grid">
    <div class="info-item"><span class="info-label">Estudiante:</span> {sol['nombre']}</div>
    <div class="info-item"><span class="info-label">Modalidad:</span> {sol['modalidad']}</div>
    <div class="info-item"><span class="info-label">Código:</span> {sol['codigo']}</div>
    <div class="info-item"><span class="info-label">IES:</span> {sol['institucion']}</div>
    <div class="info-item"><span class="info-label">Plan de estudios:</span> {sol.get('carrera_nombre', '')}</div>
    <div class="info-item"><span class="info-label">Plan externo:</span> {sol.get('plan_nombre', '')}</div>
  </div>

  <table>
    <thead>
      <tr>
        <th colspan="5" style="background:#D9E2F0;color:#1F3864;font-size:8px;">ASIGNATURAS DEL PLAN EXTERNO</th>
        <th colspan="3" style="background:#D9E2F0;color:#1F3864;font-size:8px;border-left:2px solid #fff;">CONVALIDACIÓN</th>
      </tr>
      <tr>
        <th style="width:5%;">Ciclo</th>
        <th style="width:9%;">Código</th>
        <th style="width:28%;">Nombre del curso</th>
        <th style="width:6%;">Créd.</th>
        <th style="width:12%;">Prerreq.</th>
        <th style="width:30%;">Nombre del curso</th>
        <th style="width:6%;">Créd.</th>
        <th style="width:6%;">Nota</th>
      </tr>
    </thead>
    <tbody>
"""
        for c in cursos:
            estado = c.get('estado', 'sin_validar')
            periodo = c.get('periodo_lectivo', '') or ''
            period_class = f'period-{periodo.replace("-","").replace("/","")}' if periodo and periodo in distinct_periodos else ''

            if estado == 'convalidado':
                nota = c.get('nota')
                nota_val = float(nota) if nota is not None else None
                if nota_val is not None:
                    nota_display = int(nota_val) if nota_val == int(nota_val) else nota_val
                    nota_html = f'<span style="color:#16a34a;font-weight:bold;">{nota_display}</span>' if nota_val >= 11 else f'<span style="color:#dc2626;font-weight:bold;">{nota_display}</span>'
                else:
                    nota_html = '-'
            else:
                nota_html = '-'

            # Nombre curso convalidacion segun estado
            if estado == 'convalidado':
                curso_conv = c.get('local_nombre', '')
            elif estado == 'examen_suficiencia':
                curso_conv = 'EXAMEN SUFICIENCIA'
            else:
                curso_conv = periodo if periodo else '—'

            html += f"""
      <tr class="{period_class}">
        <td class="num">{c.get('ciclo', '')}</td>
        <td class="num">{c.get('ext_codigo', '')}</td>
        <td>{c.get('ext_nombre', '')}</td>
        <td class="num">{c.get('ext_creditos', 0)}</td>
        <td class="num">{c.get('prerrequisito', '')}</td>
        <td>{curso_conv}</td>
        <td class="num">{c.get('local_creditos', 0) or ''}</td>
        <td class="num">{nota_html}</td>
      </tr>"""

        html += f"""
    </tbody>
  </table>

  <div style="text-align:right;font-weight:bold;font-size:10px;padding:6px 0;border-top:1px solid #ddd;margin-top:6px;">
    TOTAL DE CRÉDITOS DEL PROGRAMA: <span style="color:#1F3864;">{total_creditos}</span>
  </div>

  <div class="footer">
    Sistema de Convalidaciones UAI - Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}
  </div>
</div>"""

        cur.close()
        conn.close()
        return html, 200, {'Content-Type': 'text/html; charset=utf-8'}

    except Exception as e:
        cur.close()
        conn.close()
        return f'Error: {str(e)}', 500


@bp.route('/record-notas/<int:id>')
def record_notas(id):
    """Genera el record de notas del plan externo en HTML (para preview en iframe)"""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.*, COALESCE(p.apellidos_nombres,'') AS nombre,
                   COALESCE(p.dni,'') AS dni, COALESCE(p.programa,'') AS programa,
                   COALESCE(p.institucion_procedencia,'') AS institucion,
                   pe.nombre_plan AS plan_nombre, pe.periodo_academico
             FROM solicitudes s
             LEFT JOIN postulantes p ON s.postulante_id = p.id
             LEFT JOIN planes_estudio pe ON s.plan_externo_id = pe.id
             WHERE s.id = %s
        """, (id,))
        sol = cur.fetchone()
        
        if not sol:
            return '<div style="padding:2rem;text-align:center;color:#dc2626;">Solicitud no encontrada</div>', 404
        
        _oc = "CASE cp_e.ciclo WHEN 'I' THEN 1 WHEN 'II' THEN 2 WHEN 'III' THEN 3 WHEN 'IV' THEN 4 WHEN 'V' THEN 5 WHEN 'VI' THEN 6 WHEN 'VII' THEN 7 WHEN 'VIII' THEN 8 WHEN 'IX' THEN 9 WHEN 'X' THEN 10 END"
        cur.execute(f"""
            SELECT cp_e.ciclo, cp_e.codigo AS curso_codigo, 
                   cp_e.nombre_curso, cp_e.creditos,
                   COALESCE(sc.nota::TEXT, '') AS nota,
                   COALESCE(sc.estado, 'sin_validar') AS estado
            FROM cursos_plan cp_e
            LEFT JOIN solicitud_cursos sc ON sc.curso_externo_id = cp_e.id AND sc.solicitud_id = %s
            WHERE cp_e.plan_id = %s
            ORDER BY {_oc},
                     cp_e.nombre_curso
        """, (id, sol.get('plan_externo_id'),))
        cursos = cur.fetchall()
        
        cursos_por_ciclo = {}
        for c in cursos:
            ciclo = c.get('ciclo', 'X')
            if ciclo not in cursos_por_ciclo:
                cursos_por_ciclo[ciclo] = []
            cursos_por_ciclo[ciclo].append(c)
        
        total_creditos = sum(c.get('creditos', 0) for c in cursos)
        
        html = f"""
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: Arial, sans-serif; font-size: 10px; background: #fff; }}
.record-container {{
  width: 210mm; min-height: 297mm; margin: 0 auto; padding: 10mm 8mm; background: white; 
}}
.header {{ text-align: center; margin-bottom: 8px; border-bottom: 2px solid #1F3864; padding-bottom: 6px; }}
.header h1 {{ color: #1F3864; font-size: 14px; margin-bottom: 3px; }}
.header h2 {{ color: #4a7cc7; font-size: 10px; }}
.info-grid {{ display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 4px; margin-bottom: 10px; font-size: 9px; }}
.info-item {{ }}
.info-label {{ font-weight: bold; color: #555; }}
table {{ width: 100%; border-collapse: collapse; margin-bottom: 10px; font-size: 8px; }}
th {{ background: #1F3864; color: white; padding: 3px 2px; text-align: center; font-weight: 600; }}
td {{ padding: 2px; border: 1px solid #ddd; }}
tr:nth-child(even) {{ background: #f9f9f9; }}
td:nth-child(1), td:nth-child(2), td:nth-child(5), td:nth-child(6), td:nth-child(7) {{ text-align: center; }}
td:nth-child(3), td:nth-child(4) {{ text-align: left; }}
.nota-aprobado {{ color: #16a34a; font-weight: bold; }}
.nota-desaprobado {{ color: #dc2626; font-weight: bold; }}
.total {{ text-align: right; font-weight: bold; font-size: 10px; padding: 6px 0; border-top: 1px solid #ddd; margin-top: 6px; }}
.footer {{ text-align: center; font-size: 7px; color: #888; margin-top: 10px; padding-top: 6px; border-top: 1px solid #eee; }}
@page {{ size: A4; margin: 0; }}
</style>
<div class="record-container">
  <div class="header">
    <h1>RECORD DE NOTAS</h1>
    <h2>{sol['institucion'] or 'Institucion de origen'}</h2>
  </div>
  
  <div class="info-grid">
    <div class="info-item"><span class="info-label">Estudiante:</span> {sol['nombre']}</div>
    <div class="info-item"><span class="info-label">DNI:</span> {sol['dni']}</div>
    <div class="info-item"><span class="info-label">Programa:</span> {sol['programa']}</div>
    <div class="info-item"><span class="info-label">N° Solicitud:</span> {sol['codigo']}</div>
    <div class="info-item"><span class="info-label">Plan:</span> {sol['plan_nombre'] or '-'}</div>
    <div class="info-item"><span class="info-label">Periodo:</span> {sol['periodo_academico'] or '-'}</div>
  </div>
  
  <table>
    <thead>
      <tr>
        <th style="width:3%;">N°</th>
        <th style="width:5%;">Ciclo</th>
        <th style="width:8%;">Codigo</th>
        <th>Curso</th>
        <th style="width:5%;">Cred.</th>
        <th style="width:6%;">Nota</th>
      </tr>
    </thead>
    <tbody>
"""
        
        nro = 1
        for ciclo in ['I','II','III','IV','V','VI','VII','VIII','IX','X']:
            if ciclo in cursos_por_ciclo:
                for c in cursos_por_ciclo[ciclo]:
                    nota = c.get('nota')
                    nota_val = float(nota) if nota and nota != '' else None
                    if nota_val is not None:
                       nota_display = int(nota_val) if nota_val == int(nota_val) else nota_val
                       nota_html = f'<span class="nota-aprobado">{nota_display}</span>' if nota_val >= 11 else f'<span class="nota-desaprobado">{nota_display}</span>' 
                    else:
                        nota_html = '-'
                    
                    html += f"""
      <tr>
        <td>{nro}</td>
        <td>{c.get('ciclo', '')}</td>
        <td>{c.get('curso_codigo', '')}</td>
        <td>{c.get('nombre_curso', '')}</td>
        <td>{c.get('creditos', 0)}</td>
        <td>{nota_html}</td>
      </tr>"""
                    nro += 1
        
        html += f"""
    </tbody>
  </table>
  
  <div class="total">Total Creditos: <strong>{total_creditos}</strong></div>
  
  <div class="footer">
    Sistema de Convalidaciones UAI - Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}
  </div>
</div>"""
        
        cur.close()
        conn.close()
        return html, 200, {'Content-Type': 'text/html; charset=utf-8'}
        
    except Exception as e:
        cur.close()
        conn.close()
        return f'Error: {str(e)}', 500


@bp.route('/record-notas-pdf/<int:id>')
def record_notas_pdf(id):
    """Descarga el record de notas como PDF"""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT s.*, COALESCE(p.apellidos_nombres,'') AS nombre,
                   COALESCE(p.dni,'') AS dni, COALESCE(p.programa,'') AS programa,
                   COALESCE(p.institucion_procedencia,'') AS institucion,
                   pe.nombre_plan AS plan_nombre, pe.periodo_academico
             FROM solicitudes s
             LEFT JOIN postulantes p ON s.postulante_id = p.id
             LEFT JOIN planes_estudio pe ON s.plan_externo_id = pe.id
             WHERE s.id = %s
        """, (id,))
        sol = cur.fetchone()
        
        if not sol:
            return 'Solicitud no encontrada', 404
        
        _oc = "CASE cp_e.ciclo WHEN 'I' THEN 1 WHEN 'II' THEN 2 WHEN 'III' THEN 3 WHEN 'IV' THEN 4 WHEN 'V' THEN 5 WHEN 'VI' THEN 6 WHEN 'VII' THEN 7 WHEN 'VIII' THEN 8 WHEN 'IX' THEN 9 WHEN 'X' THEN 10 END"
        cur.execute(f"""
            SELECT cp_e.ciclo, cp_e.codigo AS curso_codigo, 
                   cp_e.nombre_curso, cp_e.creditos,
                   COALESCE(sc.nota::TEXT, '') AS nota
            FROM cursos_plan cp_e
            LEFT JOIN solicitud_cursos sc ON sc.curso_externo_id = cp_e.id AND sc.solicitud_id = %s
            WHERE cp_e.plan_id = %s
            ORDER BY {_oc},
                     cp_e.nombre_curso
        """, (id, sol.get('plan_externo_id')))
        cursos = cur.fetchall()
        cur.close()
        conn.close()
        
        buffer = io.BytesIO()
        # A4: 210mm x 297mm
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.6*inch, rightMargin=0.6*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=14, alignment=1, textColor=colors.HexColor('#1F3864'))
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'], fontSize=11, alignment=1, textColor=colors.HexColor('#4a7cc7'))
        normal_style = styles['Normal']
        
        elements.append(Paragraph('RECORD DE NOTAS', title_style))
        elements.append(Paragraph(sol.get('institucion') or 'Institución de Origen', subtitle_style))
        elements.append(Spacer(1, 0.3*inch))
        
        info_data = [
            ['Estudiante:', sol.get('nombre', ''), 'DNI:', sol.get('dni', '')],
            ['Programa:', sol.get('programa', ''), 'N° Solicitud:', sol.get('codigo', '')],
            ['Plan Estudios:', sol.get('plan_nombre') or '-', 'Periodo:', sol.get('periodo_academico') or '-']
        ]
        info_table = Table(info_data, colWidths=[1.2*inch, 2.3*inch, 1*inch, 2*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#555')),
            ('TEXTCOLOR', (2,0), (2,-1), colors.HexColor('#555')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.2*inch))
        
        table_data = [['N°', 'Ciclo', 'Código', 'Curso', 'Créd.', 'Nota']]
        nro = 1
        for c in cursos:
            nota = c.get('nota')
            nota_str = str(nota) if nota else '-'
            table_data.append([str(nro), c.get('ciclo', ''), c.get('curso_codigo', ''), Paragraph(c.get('nombre_curso', '')[:45], normal_style), str(c.get('creditos', 0)), nota_str])
            nro += 1
        
        cursos_table = Table(table_data, colWidths=[0.4*inch, 0.5*inch, 0.8*inch, 3.5*inch, 0.5*inch, 0.6*inch])
        cursos_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F3864')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (3,0), (3,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9f9f9')]),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(cursos_table)
        
        total_creditos = sum(c.get('creditos', 0) for c in cursos)
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(f'<b>Total Créditos:</b> {total_creditos}', normal_style))
        elements.append(Spacer(1, 0.3*inch))
        elements.append(Paragraph(f'Documento generado automáticamente - Sistema de Convalidaciones UAI - {datetime.now().strftime("%d/%m/%Y %H:%M")}', ParagraphStyle('Footer', fontSize=8, alignment=1, textColor=colors.grey)))
        
        doc.build(elements)
        buffer.seek(0)
        
        nombre_archivo = f"Record_Notas_{sol.get('nombre', 'sin_nombre').replace(' ', '_')}_{sol.get('dni', '')}.pdf"
        
        return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=nombre_archivo)
        
    except Exception as e:
        cur.close()
        conn.close()
        return f'Error: {str(e)}', 500