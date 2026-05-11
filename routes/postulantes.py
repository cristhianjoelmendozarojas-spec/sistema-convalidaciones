# routes/postulantes.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from db.conexion import get_connection
from routes.logger import registrar
from routes.generar_word import invalidar_cache
from services.postulantes_service import (
    buscar_postulantes as service_buscar_postulantes,
    get_postulantes_lista as service_get_postulantes_lista
)
from datetime import datetime
import io, os

bp_post = Blueprint('postulantes', __name__)

# ── Mapeo columnas Excel/CSV → campos BD ─────────────────────────
# Orden igual al de la plantilla Excel. Las columnas 'Estado' duplicadas
# se leen por posición en _upsert_postulante, no por nombre.
COLUMNAS = {
    'Tipo de documento':              'tipo_documento',
    'N° de DNI':                      'dni',
    'Código de estudiante':           'codigo',
    'Apellidos y nombres':            'apellidos_nombres',
    'Celular':                        'celular',
    'Correo electrónico':             'correo',
    'Departamento':                   'departamento',
    'Provincia':                      'provincia',
    'Distrito':                       'distrito',
    'Sexo':                           'sexo',
    'Fecha de nacimiento':            'fecha_nacimiento',
    'Edad':                           'edad',
    'Local':                          'local',
    'Facultad':                       'facultad',
    'Programa de estudios':           'programa',
    'Modalidad de admisión':          'modalidad_admision',
    'Semestre académico':             'semestre_academico',
    'Modalidad de estudios':          'modalidad_estudios',
    'Turno':                          'turno',
    'Asesora':                        'asesora',
    'Fecha de registro':              'fecha_registro_origen',
    'Escala de matrícula':            'escala_matricula',
    'Escala de pensiones':            'escala_pensiones',
    'Evaluación de expediente (S/.)': 'monto_expediente',
    'Fecha de pago expediente':       'fecha_pago_expediente',
    'Postulación (S/.)':              'monto_postulacion',
    'Fecha de pago postulación':      'fecha_pago_postulacion',
    'Matrícula (S/.)':                'monto_matricula',
    'Fecha de pago matrícula':        'fecha_pago_matricula',
}

# Columnas 'Estado' (3 repetidas) se leen por índice fijo
IDX_ESTADO_EXPEDIENTE  = 24   # columna 25 (base 0)
IDX_ESTADO_POSTULACION = 27   # columna 28
IDX_ESTADO_MATRICULA   = 30   # columna 31

# ── Helpers ───────────────────────────────────────────────────────

def _limpiar_fecha(valor):
    if not valor or str(valor).strip() in ('', 'nan', 'NaT', 'None'):
        return None
    s = str(valor).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None

def _limpiar_monto(valor):
    if not valor or str(valor).strip() in ('', 'nan', 'None', '—', '-'):
        return None
    try:
        # Remove currency symbols, spaces, letters except digits, dots and commas
        import re as _re
        limpio = _re.sub(r'[^0-9.,]', '', str(valor).replace(',', '.'))
        return float(limpio) if limpio else None
    except (ValueError, TypeError):
        return None

# Campos que NO se deben actualizar (claves de negocio)
CAMPOS_PROTEGIDOS = {'codigo', 'dni'}

# Etiquetas legibles para los campos
CAMPO_LABELS = {
    'tipo_documento': 'Tipo de documento', 'apellidos_nombres': 'Apellidos y nombres',
    'celular': 'Celular', 'correo': 'Correo', 'departamento': 'Departamento',
    'provincia': 'Provincia', 'distrito': 'Distrito', 'sexo': 'Sexo',
    'fecha_nacimiento': 'Fecha nacimiento', 'edad': 'Edad', 'local': 'Local',
    'facultad': 'Facultad', 'programa': 'Programa', 'modalidad_admision': 'Modalidad admisión',
    'semestre_academico': 'Semestre académico', 'modalidad_estudios': 'Modalidad estudios',
    'turno': 'Turno', 'asesora': 'Asesora', 'fecha_registro_origen': 'Fecha registro',
    'escala_matricula': 'Escala matrícula', 'escala_pensiones': 'Escala pensiones',
    'monto_expediente': 'Monto expediente', 'estado_expediente': 'Estado expediente',
    'fecha_pago_expediente': 'Fecha pago expediente', 'monto_postulacion': 'Monto postulación',
    'estado_postulacion': 'Estado postulación', 'fecha_pago_postulacion': 'Fecha pago postulación',
    'monto_matricula': 'Monto matrícula', 'estado_matricula': 'Estado matrícula',
    'fecha_pago_matricula': 'Fecha pago matrícula',
}

# Campos de texto que se normalizan a MAYÚSCULAS al comparar
_CAMPOS_UPPER = {
    'apellidos_nombres', 'programa', 'facultad', 'departamento',
    'provincia', 'distrito', 'local', 'institucion_procedencia',
}

def _norm_texto(v: str) -> str:
    """Normaliza texto: strip, colapsa espacios múltiples."""
    return ' '.join(v.upper().split()) if v else ''

def _norm_valor(campo: str, raw: str, vacios: set) -> str:
    """Devuelve el valor normalizado según el tipo de campo, o '' si está vacío."""
    v = raw.strip() if raw else ''
    if v in vacios:
        return ''
    if campo in _CAMPOS_MONTO:
        n = _limpiar_monto(v)
        return f'{n:.2f}' if n is not None else ''
    if campo in _CAMPOS_FECHA:
        f = _limpiar_fecha(v)
        return f if f else ''
    if campo in _CAMPOS_INT:
        try: return str(int(float(v)))
        except: return ''
    if campo in _CAMPOS_UPPER:
        return _norm_texto(v)
    # Texto genérico: strip y colapsar espacios
    return ' '.join(v.split())


def _detectar_cambios(cur, codigo: str, fila_nueva: dict) -> list:
    """Compara valores actuales en BD con los del archivo. Retorna solo diffs reales."""
    from datetime import date as _date, datetime as _datetime

    cur.execute("SELECT * FROM postulantes WHERE codigo=%s LIMIT 1", (codigo,))
    actual = cur.fetchone()
    if not actual:
        return []

    cambios = []
    VACIOS = {'', 'None', 'nan', 'NaT', 'none', 'null',
              '0000-00-00', '0000-00-00 00:00:00', 'NULL'}

    for campo, val_nuevo in fila_nueva.items():
        if campo in CAMPOS_PROTEGIDOS:
            continue

        val_actual = actual.get(campo)

        # Convertir BD date/datetime a string ISO
        if isinstance(val_actual, (_date, _datetime)):
            va_raw = val_actual.strftime('%Y-%m-%d')
        elif val_actual is None:
            va_raw = ''
        else:
            va_raw = str(val_actual).strip()

        vn_raw = str(val_nuevo).strip() if val_nuevo is not None else ''

        # Normalizar ambos con la misma función
        va_norm = _norm_valor(campo, va_raw, VACIOS)
        vn_norm = _norm_valor(campo, vn_raw, VACIOS)

        # Solo registrar si hay diferencia real y el nuevo no está vacío
        if va_norm != vn_norm and vn_norm != '':
            cambios.append({
                'campo':    campo,
                'label':    CAMPO_LABELS.get(campo, campo),
                'anterior': va_norm if va_norm else '—',
                'nuevo':    vn_norm,
            })
    return cambios

def _upsert_postulante(cur, fila: dict, fila_raw: tuple = None):
    """INSERT si es nuevo, UPDATE si el código ya existe. Retorna 'nuevo' o 'actualizado'.
    fila_raw es la fila original (tupla) para leer columnas 'Estado' duplicadas por índice.
    """
    # Leer los 3 campos 'Estado' por índice fijo (no por nombre, están duplicados)
    def _estado(idx):
        if fila_raw and idx < len(fila_raw):
            v = fila_raw[idx]
            return str(v).strip() if v else ''
        return ''

    sql = """
        INSERT INTO postulantes
            (codigo,tipo_documento,dni,apellidos_nombres,celular,correo,
             departamento,provincia,distrito,sexo,fecha_nacimiento,edad,
             local,facultad,programa,modalidad_admision,semestre_academico,
             modalidad_estudios,turno,asesora,fecha_registro_origen,
             escala_matricula,escala_pensiones,
             monto_expediente,estado_expediente,fecha_pago_expediente,
             monto_postulacion,estado_postulacion,fecha_pago_postulacion,
             monto_matricula,estado_matricula,fecha_pago_matricula)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (codigo) DO UPDATE SET
            tipo_documento=EXCLUDED.tipo_documento,
            apellidos_nombres=EXCLUDED.apellidos_nombres, celular=EXCLUDED.celular,
            correo=EXCLUDED.correo, departamento=EXCLUDED.departamento,
            provincia=EXCLUDED.provincia, distrito=EXCLUDED.distrito,
            sexo=EXCLUDED.sexo, fecha_nacimiento=EXCLUDED.fecha_nacimiento,
            edad=EXCLUDED.edad, local=EXCLUDED.local, facultad=EXCLUDED.facultad,
            programa=EXCLUDED.programa, modalidad_admision=EXCLUDED.modalidad_admision,
            semestre_academico=EXCLUDED.semestre_academico,
            modalidad_estudios=EXCLUDED.modalidad_estudios, turno=EXCLUDED.turno,
            asesora=EXCLUDED.asesora, fecha_registro_origen=EXCLUDED.fecha_registro_origen,
            escala_matricula=EXCLUDED.escala_matricula, escala_pensiones=EXCLUDED.escala_pensiones,
            monto_expediente=EXCLUDED.monto_expediente, estado_expediente=EXCLUDED.estado_expediente,
            fecha_pago_expediente=EXCLUDED.fecha_pago_expediente,
            monto_postulacion=EXCLUDED.monto_postulacion, estado_postulacion=EXCLUDED.estado_postulacion,
            fecha_pago_postulacion=EXCLUDED.fecha_pago_postulacion,
            monto_matricula=EXCLUDED.monto_matricula, estado_matricula=EXCLUDED.estado_matricula,
            fecha_pago_matricula=EXCLUDED.fecha_pago_matricula,
            fecha_actualizacion=NOW()
    """
    edad = fila.get('edad')
    try: edad = int(str(edad).strip()) if edad and str(edad).strip() not in ('','nan','None') else None
    except: edad = None

    valores = (
        str(fila.get('codigo','')).strip().upper(),
        fila.get('tipo_documento',''), fila.get('dni',''),
        str(fila.get('apellidos_nombres','')).strip().upper(),
        fila.get('celular',''), fila.get('correo',''),
        fila.get('departamento',''), fila.get('provincia',''), fila.get('distrito',''),
        fila.get('sexo',''), _limpiar_fecha(fila.get('fecha_nacimiento')), edad,
        fila.get('local',''), fila.get('facultad',''),
        str(fila.get('programa','')).strip().upper(),
        fila.get('modalidad_admision',''), fila.get('semestre_academico',''),
        fila.get('modalidad_estudios',''), fila.get('turno',''), fila.get('asesora',''),
        _limpiar_fecha(fila.get('fecha_registro_origen')),
        fila.get('escala_matricula',''), fila.get('escala_pensiones',''),
        _limpiar_monto(fila.get('monto_expediente')), _estado(IDX_ESTADO_EXPEDIENTE),
        _limpiar_fecha(fila.get('fecha_pago_expediente')),
        _limpiar_monto(fila.get('monto_postulacion')), _estado(IDX_ESTADO_POSTULACION),
        _limpiar_fecha(fila.get('fecha_pago_postulacion')),
        _limpiar_monto(fila.get('monto_matricula')), _estado(IDX_ESTADO_MATRICULA),
        _limpiar_fecha(fila.get('fecha_pago_matricula')),
    )
    cur.execute(sql, valores)
    return 'nuevo' if cur.rowcount == 1 else 'actualizado'

def _leer_archivo(archivo, ext):
    """Lee Excel o CSV y retorna (encabezados, filas)."""
    if ext in ('.xlsx', '.xls'):
        import openpyxl
        wb  = openpyxl.load_workbook(archivo, data_only=True)
        ws  = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, None
        return [str(c).strip() if c else '' for c in rows[0]], rows[1:]
    else:
        import csv, io as _io
        contenido = archivo.read().decode('utf-8-sig')
        reader    = csv.reader(_io.StringIO(contenido))
        rows      = list(reader)
        if not rows:
            return None, None
        return [c.strip() for c in rows[0]], rows[1:]

# ── Rutas ─────────────────────────────────────────────────────────

@bp_post.route('/')
def index():
    page = max(1, int(request.args.get('page', 1)))
    per_page = min(50, max(10, int(request.args.get('per_page', 20))))
    offset = (page - 1) * per_page
    
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    q = request.args.get('q','').strip()
    programa  = request.args.get('programa','').strip()
    modalidad = request.args.get('modalidad','').strip()
    estado    = request.args.get('estado','').strip()

    where, params = ['1=1'], []
    if q:
        where.append('(codigo LIKE %s OR apellidos_nombres LIKE %s OR dni LIKE %s)')
        like = f'%{q}%'; params += [like, like, like]
    if programa:
        where.append('programa=%s'); params.append(programa)
    if modalidad:
        where.append('modalidad_estudios=%s'); params.append(modalidad)

    cur.execute(f"SELECT COUNT(*) as total FROM postulantes p WHERE {' AND '.join(where)}", params)
    total = cur.fetchone()['total']
    total_pages = (total + per_page - 1) // per_page if total > 0 else 1

    cur.execute(f"""
        SELECT p.*,
               s.id AS solicitud_id,
               s.codigo AS solicitud_codigo,
               s.estado AS solicitud_estado
        FROM postulantes p
        LEFT JOIN solicitudes s ON s.postulante_id = p.id
        WHERE {' AND '.join(where)}
        ORDER BY p.apellidos_nombres
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])
    postulantes = cur.fetchall()

    if estado:
        if estado == 'emitido':
            postulantes = [p for p in postulantes if p.get('solicitud_estado') == 'emitido']
        elif estado == 'en_proceso':
            postulantes = [p for p in postulantes if p.get('solicitud_estado') in ('borrador', None)]
        elif estado == 'sin_solicitud':
            postulantes = [p for p in postulantes if not p.get('solicitud_id')]

    cur.execute("SELECT DISTINCT programa FROM postulantes WHERE programa IS NOT NULL ORDER BY programa")
    programas = [r['programa'] for r in cur.fetchall()]
    cur.execute("SELECT DISTINCT modalidad_estudios FROM postulantes WHERE modalidad_estudios IS NOT NULL ORDER BY modalidad_estudios")
    modalidades = [r['modalidad_estudios'] for r in cur.fetchall()]

    cur.execute("""
        SELECT codigo, apellidos_nombres, fecha_actualizacion
        FROM postulantes
        WHERE fecha_actualizacion >= NOW() - INTERVAL '24 HOURS'
          AND fecha_actualizacion != fecha_importacion
        ORDER BY fecha_actualizacion DESC
        LIMIT 100
    """)
    duplicados_recientes = cur.fetchall()

    cur.close(); conn.close()
    return render_template('postulantes/lista.html',
        postulantes=postulantes, programas=programas, modalidades=modalidades,
        q=q, programa_sel=programa, modalidad_sel=modalidad, estado_sel=estado,
        duplicados_recientes=duplicados_recientes,
        page=page, per_page=per_page, total=total, total_pages=total_pages)


@bp_post.route('/importar', methods=['GET','POST'])
def importar():
    if request.method == 'GET':
        # Leer y limpiar pendientes de sesión
        pendientes = session.pop('pendientes_importar', None)
        session.modified = True
        return render_template('postulantes/importar.html',
                               pendientes_session=pendientes)

    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        flash('Selecciona un archivo.', 'danger')
        return redirect(url_for('postulantes.importar'))

    ext = os.path.splitext(archivo.filename)[1].lower()
    if ext not in ('.xlsx', '.xls', '.csv'):
        flash('Formato no válido. Usa .xlsx, .xls o .csv', 'danger')
        return redirect(url_for('postulantes.importar'))

    try:
        encabezados, datos = _leer_archivo(archivo, ext)
        if encabezados is None:
            flash('El archivo está vacío.', 'danger')
            return redirect(url_for('postulantes.importar'))

        col_map = {}
        for nombre_col, campo_bd in COLUMNAS.items():
            try: col_map[campo_bd] = encabezados.index(nombre_col)
            except ValueError: pass

        if 'codigo' not in col_map:
            flash('El archivo no tiene la columna "Código de estudiante".', 'danger')
            return redirect(url_for('postulantes.importar'))

        conn = get_connection()
        cur     = conn.cursor(dictionary=True)   # para SELECT
        cur_cmp = conn.cursor(dictionary=True)   # cursor dedicado para comparar
        nuevos = errores = 0
        pendientes = []
        filas_nuevas = []

        for fila in datos:
            try:
                fila_dict = {campo: (fila[idx] if idx < len(fila) else None)
                             for campo, idx in col_map.items()}
                codigo = str(fila_dict.get('codigo') or '').strip().upper()
                if not codigo:
                    continue

                # Verificar si existe
                cur.execute("SELECT id FROM postulantes WHERE codigo=%s LIMIT 1", (codigo,))
                existe = cur.fetchone()

                if not existe:
                    filas_nuevas.append((fila_dict, fila))
                    nuevos += 1
                else:
                    # Cursor dedicado para evitar conflictos de estado
                    cambios = _detectar_cambios(cur_cmp, codigo, fila_dict)
                    if cambios:
                        pendientes.append({
                            'codigo':    codigo,
                            'nombre':    str(fila_dict.get('apellidos_nombres','')).strip().upper(),
                            'cambios':   cambios,
                            'fila_dict': {k: (str(v) if v is not None else '') for k,v in fila_dict.items()},
                        })
            except Exception as e:
                errores += 1
                print(f'ERROR fila: {e}', flush=True)

        # Insertar nuevos directamente
        cur2 = conn.cursor()
        for fd, fr in filas_nuevas:
            try:
                _upsert_postulante(cur2, fd, fila_raw=fr)
            except Exception as e:
                print(f'ERROR insert: {e}', flush=True)

        conn.commit()
        cur.close(); cur2.close(); conn.close()

        msg = f'Importación: {nuevos} nuevo(s)'
        if errores:
            msg += f', {errores} con error'

        # Si no hay cambios pendientes → ir directo a postulantes
        if not pendientes:
            flash(msg, 'success')
            return redirect(url_for('postulantes.index'))

        msg += f', {len(pendientes)} con cambios pendientes de revisión'
        flash(msg, 'warning')
        return render_template('postulantes/importar.html',
                               pendientes_session=pendientes)

    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
        return redirect(url_for('postulantes.importar'))


# Campos por tipo para normalización al guardar
_CAMPOS_MONTO  = {'monto_expediente','monto_postulacion','monto_matricula'}
_CAMPOS_FECHA  = {'fecha_nacimiento','fecha_registro_origen','fecha_pago_expediente',
                  'fecha_pago_postulacion','fecha_pago_matricula'}
_CAMPOS_INT    = {'edad'}

def _normalizar_para_bd(campo: str, valor: str):
    """Convierte el valor string al tipo correcto según el campo de BD."""
    v = str(valor).strip() if valor is not None else ''
    if v in ('', '—', '-', 'None', 'nan', 'NaT'):
        return None
    if campo in _CAMPOS_MONTO:
        return _limpiar_monto(v)
    if campo in _CAMPOS_FECHA:
        return _limpiar_fecha(v)
    if campo in _CAMPOS_INT:
        try: return int(float(v))
        except: return None
    return v or None


@bp_post.route('/aprobar-cambios', methods=['POST'])
def aprobar_cambios():
    """Aplica los cambios aprobados (todos o seleccionados)."""
    data = request.get_json()
    aprobados = data.get('aprobados', [])

    if not aprobados:
        # Descartar — solo limpiar sesión
        session.pop('pendientes_importar', None)
        return jsonify({'ok': True, 'actualizados': 0})

    conn = get_connection(); cur = conn.cursor()
    try:
        por_codigo = {}
        for item in aprobados:
            c = item['codigo']
            if c not in por_codigo:
                por_codigo[c] = {}
            campo = item['campo']
            if campo not in CAMPOS_PROTEGIDOS:
                # Normalizar al tipo correcto antes de guardar
                por_codigo[c][campo] = _normalizar_para_bd(campo, item['valor'])

        for codigo, campos in por_codigo.items():
            if not campos:
                continue
            sets = ', '.join(f"{k}=%s" for k in campos)
            vals = list(campos.values()) + [codigo]
            cur.execute(
                f"UPDATE postulantes SET {sets}, fecha_actualizacion=NOW() WHERE codigo=%s",
                vals
            )

        conn.commit()
        session.pop('pendientes_importar', None)
        return jsonify({'ok': True, 'actualizados': len(por_codigo)})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)})
    finally:
        cur.close(); conn.close()


@bp_post.route('/plantilla-excel')
def descargar_plantilla():
    """Genera y descarga un Excel con encabezados centrados y fila de ejemplo alineada a la derecha."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Postulantes'

    borde = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # Encabezados reales del archivo (con los 'Estado' duplicados en posición correcta)
    encabezados_excel = [
        'Tipo de documento', 'N° de DNI', 'Código de estudiante', 'Apellidos y nombres',
        'Celular', 'Correo electrónico', 'Departamento', 'Provincia', 'Distrito',
        'Sexo', 'Fecha de nacimiento', 'Edad', 'Local', 'Facultad',
        'Programa de estudios', 'Modalidad de admisión', 'Semestre académico',
        'Modalidad de estudios', 'Turno', 'Asesora', 'Fecha de registro',
        'Escala de matrícula', 'Escala de pensiones',
        'Evaluación de expediente (S/.)', 'Estado', 'Fecha de pago expediente',
        'Postulación (S/.)', 'Estado', 'Fecha de pago postulación',
        'Matrícula (S/.)', 'Estado', 'Fecha de pago matrícula',
    ]

    for col_idx, titulo in enumerate(encabezados_excel, 1):
        celda = ws.cell(row=1, column=col_idx, value=titulo)
        celda.font      = Font(bold=True, color='FFFFFF', size=10)
        celda.fill      = PatternFill('solid', fgColor='0C1D3A')
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.border    = borde
        ws.column_dimensions[celda.column_letter].width = max(14, len(titulo) * 1.05)

    ws.row_dimensions[1].height = 35

    # Fila de ejemplo con datos alineados a la derecha
    ejemplo = [
        'DNI', '12345678', 'EST-001', 'APELLIDOS NOMBRES',
        '987654321', 'correo@example.com', 'ICA', 'ICA', 'ICA',
        'F', '01/01/2000', '24', 'CHINCHA', 'CIENCIAS DE LA SALUD',
        'ENFERMERÍA', 'Graduados y Titulados', '2026-I',
        'Presencial', 'Noche', 'ANA GARCIA', '14/03/2026',
        'A', 'A',
        '150.00', 'PAGADO', '14/03/2026',
        '100.00', 'PAGADO', '14/03/2026',
        '200.00', 'PENDIENTE', '',
    ]
    for col_idx, valor in enumerate(ejemplo, 1):
        celda = ws.cell(row=2, column=col_idx, value=valor)
        celda.alignment = Alignment(horizontal='right', vertical='center')
        celda.font      = Font(color='888888', italic=True, size=9)
        celda.border    = borde

    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A2'  # congela la fila de encabezados

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='plantilla_postulantes.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp_post.route('/eliminar/<int:id>', methods=['POST'])
def eliminar(id):
    data = request.get_json() or {}
    force = data.get('force', False)
    
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    try:
        cur.execute("SELECT id, codigo, estado FROM solicitudes WHERE postulante_id=%s", (id,))
        solicitudes = cur.fetchall()
        
        tiene_emitidas = any(s['estado'] == 'emitido' for s in solicitudes)
        codigos = ', '.join([s['codigo'] for s in solicitudes])
        
        if solicitudes and not force:
            return jsonify({
                'ok': False,
                'error': 'El postulante tiene solicitudes asociadas',
                'solicitudes': codigos,
                'tiene_emitidas': tiene_emitidas,
                'requiere_confirmacion': True
            }), 400
        
        if tiene_emitidas and not force:
            return jsonify({
                'ok': False,
                'error': 'El postulante tiene solicitudes EMITIDAS. ¿Deseas forzar la eliminación?',
                'solicitudes': codigos,
                'tiene_emitidas': True,
                'requiere_confirmacion': True
            }), 400
        
        if force:
            for s in solicitudes:
                cur.execute("DELETE FROM solicitud_cursos WHERE solicitud_id=%s", (s['id'],))
                cur.execute("DELETE FROM solicitudes WHERE id=%s", (s['id'],))
            registrar('eliminar', 'postulantes', f'Postulante eliminado (FORZADO) con {len(solicitudes)} solicitud(es): {codigos}', id)
        
        cur.execute("DELETE FROM postulantes WHERE id=%s", (id,))
        conn.commit()
        if not force:
            registrar('eliminar', 'postulantes', f'Postulante eliminado: id={id}', id)
        return jsonify({'ok': True, 'force': force})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        cur.close(); conn.close()


@bp_post.route('/ver/<int:id>')
def ver(id):
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM postulantes WHERE id=%s", (id,))
    p = cur.fetchone()
    if not p:
        flash('Postulante no encontrado.', 'danger')
        return redirect(url_for('postulantes.index'))

    # Crear items del checklist si no existen aún
    _asegurar_checklist_docs(cur, p)
    conn.commit()

    cur.execute("SELECT * FROM checklist_documentos WHERE postulante_id=%s ORDER BY id", (id,))
    items = cur.fetchall()
    total = len(items)
    entregados = sum(1 for i in items if i['entregado'])

    cur.close(); conn.close()
    return render_template('postulantes/detalle.html', p=p, items=items, total=total, entregados=entregados)


def _asegurar_checklist_docs(cur, postulante):
    """Crea los 20 documentos base en checklist_documentos si aún no existen."""
    cur.execute("SELECT documento FROM checklist_documentos WHERE postulante_id=%s", (postulante['id'],))
    existentes = {r['documento'] for r in cur.fetchall()}
    for nombre, es_silabo in DOCUMENTOS_BASE:
        if nombre not in existentes:
            cur.execute(
                "INSERT INTO checklist_documentos (postulante_id, documento, es_silabo) VALUES (%s,%s,%s)",
                (postulante['id'], nombre, bool(es_silabo))
            )


# ── Checklist de documentos ───────────────────────────────────────

DOCUMENTOS_BASE = [
    ('Formulario de inscripción',                                  False),
    ('Ficha de registro',                                          False),
    ('Certificado de estudios secundarios',                        False),
    ('Fotografía tamaño carnet',                                   False),
    ('Constancia de egresado',                                     False),
    ('Copia del título',                                           False),
    ('Documento apostillado (solo estudios en el extranjero)',     False),
    ('Constancia de egreso o título de estudios policiales',       False),
    ('Documento de identidad',                                     False),
    ('01 fotografía tamaño carnet',                                False),
    ('DECLARACIÓN JURADA F-01/DJ ADMISIÓN',                        False),
    ('DECLARACIÓN JURADA F-02/DJ ADMISIÓN',                        False),
    ('DECLARACIÓN JURADA F-03/DJ CONVALIDACIÓN',                   False),
    ('DECLARACIÓN JURADA F-04/DJ DE VERACIDAD DE DOCUMENTOS',      False),
    ('DECLARACIÓN JURADA F-05/DJ VALIDACIÓN DE ESTUDIOS',          False),
    ('DECLARACIÓN JURADA FCS-01/DJ INGRESANTE FCS',                False),
    ('Certificado de estudios superiores',                         False),
    ('Constancia de ingreso',                                      False),
    ('Constancia de primera matrícula',                            False),
    ('Constancia de buena conducta',                               False),
    ('Sílabos fedateados',                                         True),   # es_silabo=True
]

@bp_post.route('/ver/<int:id>/checklist')
def checklist(id):
    conn = get_connection(); cur = conn.cursor(dictionary=True)

    cur.execute("SELECT id, apellidos_nombres, programa FROM postulantes WHERE id=%s", (id,))
    p = cur.fetchone()
    if not p:
        flash('Postulante no encontrado.', 'danger')
        return redirect(url_for('postulantes.index'))

    # Crear items faltantes si aún no existen
    cur.execute("SELECT documento FROM checklist_documentos WHERE postulante_id=%s", (id,))
    existentes = {r['documento'] for r in cur.fetchall()}

    for nombre, es_silabo in DOCUMENTOS_BASE:
        if nombre not in existentes:
            cur.execute("""
                INSERT INTO checklist_documentos (postulante_id, documento, es_silabo)
                VALUES (%s, %s, %s)
            """, (id, nombre, bool(es_silabo)))
    conn.commit()

    cur.execute("""
        SELECT * FROM checklist_documentos
        WHERE postulante_id=%s ORDER BY id
    """, (id,))
    items = cur.fetchall()
    total = len(items)
    entregados = sum(1 for i in items if i['entregado'])

    cur.close(); conn.close()
    return render_template('postulantes/checklist.html',
        p=p, items=items, total=total, entregados=entregados)


@bp_post.route('/checklist/guardar/<int:item_id>', methods=['POST'])
def guardar_item(item_id):
    d = request.get_json()
    entregado = True if d.get('entregado') else False
    conn = get_connection(); cur = conn.cursor()
    try:
        if 'archivo' in d:
            cur.execute("UPDATE checklist_documentos SET archivo=NULL WHERE id=%s", (item_id,))
            conn.commit(); return jsonify({'ok': True})
        elif entregado:
            cur.execute("""
                UPDATE checklist_documentos
                SET entregado=%s, tipo_doc=%s, detalle=%s, observacion=%s, fecha_entrega=%s
                WHERE id=%s
            """, (entregado, d.get('tipo_doc',''), d.get('detalle',''),
                    d.get('observacion',''), d.get('fecha_entrega') or None, item_id))
        else:
            cur.execute("""
                UPDATE checklist_documentos
                SET entregado=FALSE, fecha_entrega=NULL
                WHERE id=%s
            """, (item_id,))
        conn.commit(); return jsonify({'ok': True})
    except Exception as e:
        conn.rollback(); return jsonify({'ok': False, 'error': str(e)}), 500
    finally: cur.close(); conn.close()


@bp_post.route('/checklist/agregar/<int:postulante_id>', methods=['POST'])
def agregar_item(postulante_id):
    """Agregar documento personalizado (ej: sílabo con detalle específico)."""
    d = request.get_json()
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO checklist_documentos (postulante_id, documento, es_silabo, detalle)
            VALUES (%s, %s, %s, %s)
        """, (postulante_id, d.get('documento',''), bool(d.get('es_silabo', False)), d.get('detalle','')))
        conn.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})
    except Exception as e:
        conn.rollback(); return jsonify({'ok': False, 'error': str(e)}), 500
    finally: cur.close(); conn.close()


@bp_post.route('/checklist/eliminar/<int:item_id>', methods=['POST'])
def eliminar_item(item_id):
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("DELETE FROM checklist_documentos WHERE id=%s", (item_id,))
        conn.commit(); return jsonify({'ok': True})
    except Exception as e:
        conn.rollback(); return jsonify({'ok': False, 'error': str(e)}), 500
    finally: cur.close(); conn.close()


@bp_post.route('/checklist/upload/<int:item_id>', methods=['POST'])
def upload_item_file(item_id):
    """Sube archivo para un item del checklist y actualiza el registro."""
    import uuid

    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No se encontró archivo'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'error': 'No se seleccionó archivo'}), 400

    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else 'jpg'
    if ext not in {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf', 'doc', 'docx', 'xls', 'xlsx', 'csv'}:
        return jsonify({'ok': False, 'error': 'Tipo de archivo no permitido'}), 400

    try:
        folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'uploads')
        os.makedirs(folder, exist_ok=True)

        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("SELECT postulante_id FROM checklist_documentos WHERE id=%s", (item_id,))
        row = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return jsonify({'ok': False, 'error': 'Item no encontrado'}), 404

        filename = f"item_{item_id}_{uuid.uuid4().hex[:8]}.{ext}"
        filepath = os.path.join(folder, filename)
        file.save(filepath)

        cur.execute("""
            UPDATE checklist_documentos
            SET archivo=%s
            WHERE id=%s
        """, (filename, item_id))
        conn.commit()

        # Obtener el item actualizado para devolver estado completo
        cur.execute("""
            SELECT id, documento, archivo, es_silabo, detalle, tipo_doc, entregado, observacion, fecha_entrega
            FROM checklist_documentos WHERE id=%s
        """, (item_id,))
        item_actualizado = cur.fetchone()
        cur.close(); conn.close()

        return jsonify({
            'ok': True,
            'filename': filename,
            'url': f"/uploads/{filename}",
            'es_imagen': ext in {'png', 'jpg', 'jpeg', 'gif', 'webp'},
            'item': item_actualizado
        })
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp_post.route('/checklist/item/<int:item_id>')
def get_checklist_item(item_id):
    """Obtiene un item individual del checklist."""
    conn = get_connection()
    cur = conn.cursor(dictionary=True)
    cur.execute("""
        SELECT id, documento, archivo, es_silabo, detalle, tipo_doc, entregado, observacion, fecha_entrega
        FROM checklist_documentos WHERE id=%s
    """, (item_id,))
    item = cur.fetchone()
    cur.close(); conn.close()
    if not item:
        return jsonify({'ok': False, 'error': 'No encontrado'}), 404
    return jsonify({'ok': True, 'item': item})


# ── RECEPCIONES DE DOCUMENTOS ─────────────────────────────────────

@bp_post.route('/checklist/recepciones/<int:documento_id>')
def get_recepciones(documento_id):
    """Obtiene el historial de recepciones de un documento."""
    try:
        conn = get_connection()
        cur = conn.cursor(dictionary=True)
        cur.execute("""
            SELECT id, area, fecha_recepcion, observaciones, registrado_por, created_at
            FROM checklist_recepciones
            WHERE documento_id = %s
            ORDER BY fecha_recepcion DESC, id DESC
        """, (documento_id,))
        recepciones = cur.fetchall()
        cur.close(); conn.close()
        return jsonify({'ok': True, 'recepciones': recepciones})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp_post.route('/checklist/recepciones', methods=['POST'])
def registrar_recepcion():
    """Registra una nueva recepcion de documento."""
    d = request.get_json()
    documento_id = d.get('documento_id')
    area = d.get('area', 'Oficina').strip()
    fecha = d.get('fecha_recepcion')
    observaciones = d.get('observaciones', '').strip()
    
    if not documento_id or not fecha:
        return jsonify({'ok': False, 'error': 'Documento y fecha requeridos'}), 400
    
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO checklist_recepciones (documento_id, area, fecha_recepcion, observaciones, registrado_por)
            VALUES (%s, %s, %s, %s, %s)
        """, (documento_id, area, fecha, observaciones or None, session.get('usuario_nombre')))
        conn.commit()
        recepcion_id = cur.lastrowid
        cur.close(); conn.close()
        return jsonify({'ok': True, 'id': recepcion_id})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@bp_post.route('/checklist/recepciones/<int:recepcion_id>', methods=['DELETE'])
def eliminar_recepcion(recepcion_id):
    """Elimina un registro de recepcion."""
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("DELETE FROM checklist_recepciones WHERE id=%s", (recepcion_id,))
        conn.commit()
        cur.close(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── EDITAR POSTULANTE ──────────────────────────────────────────────

CAMPOS_EDITABLES = {
    'tipo_documento', 'dni', 'apellidos_nombres', 'celular', 'correo',
    'departamento', 'provincia', 'distrito', 'sexo', 'fecha_nacimiento', 'edad',
    'local', 'facultad', 'programa', 'modalidad_admision', 'semestre_academico',
    'modalidad_estudios', 'turno', 'asesora', 'escala_matricula', 'escala_pensiones'
}


@bp_post.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    conn = get_connection(); cur = conn.cursor(dictionary=True)
    
    if request.method == 'GET':
        cur.execute("SELECT * FROM postulantes WHERE id=%s", (id,))
        p = cur.fetchone()
        cur.close(); conn.close()
        if not p:
            flash('Postulante no encontrado.', 'danger')
            return redirect(url_for('postulantes.index'))
        return render_template('postulantes/editar.html', p=p)
    
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': 'Datos requeridos'}), 400
    
    updates = []
    valores = []
    for campo, valor in data.items():
        if campo in CAMPOS_EDITABLES:
            updates.append(f"{campo}=%s")
            valores.append(valor if valor else None)
    
    if not updates:
        return jsonify({'ok': False, 'error': 'No hay campos válidos para actualizar'})
    
    updates.append("fecha_actualizacion=NOW()")
    valores.append(id)
    
    try:
        cur.execute(f"UPDATE postulantes SET {', '.join(updates)} WHERE id=%s", valores)
        conn.commit()
        
        # Invalidar cache de solicitudes relacionadas con este postulado
        cur.execute("SELECT id FROM solicitudes WHERE postulante_id=%s", (id,))
        for row in cur.fetchall():
            invalidar_cache(row[0])
        
        registrar('editar', 'postulantes', f'Datos actualizados para postulado id={id}', id)
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        cur.close(); conn.close()


@bp_post.route('/actualizar-pago/<int:id>', methods=['POST'])
def actualizar_pago(id):
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': 'Datos requeridos'}), 400
    
    tipo = data.get('tipo')
    monto = data.get('monto')
    estado = data.get('estado')
    fecha = data.get('fecha')
    
    campos = {
        'expediente': ('monto_expediente', 'estado_expediente', 'fecha_pago_expediente'),
        'postulacion': ('monto_postulacion', 'estado_postulacion', 'fecha_pago_postulacion'),
        'matricula': ('monto_matricula', 'estado_matricula', 'fecha_pago_matricula'),
    }
    
    if tipo not in campos:
        return jsonify({'ok': False, 'error': 'Tipo de pago inválido'}), 400
    
    col_monto, col_estado, col_fecha = campos[tipo]
    
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute(f"""
            UPDATE postulantes 
            SET {col_monto}=%s, {col_estado}=%s, {col_fecha}=%s, fecha_actualizacion=NOW()
            WHERE id=%s
        """, (float(monto) if monto else None, estado, fecha, id))
        conn.commit()
        registrar('editar', 'postulantes', f'Pago {tipo} actualizado para postulante id={id}', id)
        return jsonify({'ok': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        cur.close(); conn.close()


# ── CREAR POSTULANTE ──────────────────────────────────────────────

CAMPOS_CREAR = [
    'codigo', 'tipo_documento', 'dni', 'apellidos_nombres',
    'celular', 'correo', 'departamento', 'provincia', 'distrito',
    'sexo', 'fecha_nacimiento', 'edad', 'local', 'facultad', 'programa',
    'modalidad_admision', 'semestre_academico', 'modalidad_estudios',
    'turno', 'asesora', 'fecha_registro_origen',
    'escala_matricula', 'escala_pensiones',
    'monto_expediente', 'fecha_pago_expediente',
    'monto_postulacion', 'fecha_pago_postulacion',
    'monto_matricula', 'fecha_pago_matricula'
]


@bp_post.route('/crear', methods=['POST'])
def crear():
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'error': 'Datos requeridos'}), 400
    
    codigo = data.get('codigo', '').strip().upper()
    dni = data.get('dni', '').strip()
    
    if not codigo or not dni:
        return jsonify({'ok': False, 'error': 'Código y DNI son requeridos'}), 400
    
    conn = get_connection(); cur = conn.cursor()
    try:
        cur.execute("SELECT id FROM postulantes WHERE codigo=%s LIMIT 1", (codigo,))
        if cur.fetchone():
            return jsonify({'ok': False, 'error': f'El código {codigo} ya existe'}), 400
        
        cur.execute("SELECT id FROM postulantes WHERE dni=%s LIMIT 1", (dni,))
        if cur.fetchone():
            return jsonify({'ok': False, 'error': f'El DNI {dni} ya está registrado'}), 400
        
        campos = []
        valores = []
        campos_monto = {'monto_expediente', 'monto_postulacion', 'monto_matricula'}
        
        for campo in CAMPOS_CREAR:
            if campo in data and data[campo] and str(data[campo]).strip():
                valor = data[campo]
                if campo in campos_monto:
                    try:
                        valor = float(valor)
                    except (ValueError, TypeError):
                        valor = None
                campos.append(campo)
                valores.append(valor)
        
        if campos:
            sql = f"INSERT INTO postulantes ({','.join(campos)}, fecha_importacion) VALUES ({','.join(['%s']*len(campos))}, NOW())"
            cur.execute(sql, valores)
        else:
            sql = "INSERT INTO postulantes (codigo, dni, apellidos_nombres, fecha_importacion) VALUES (%s, %s, %s, NOW())"
            cur.execute(sql, (codigo, dni, 'SIN NOMBRE'))
        
        conn.commit()
        nuevo_id = cur.lastrowid
        registrar('crear', 'postulantes', f'Postulante creado: {codigo}', nuevo_id)
        return jsonify({'ok': True, 'id': nuevo_id})
    except Exception as e:
        conn.rollback()
        return jsonify({'ok': False, 'error': str(e)}), 500
    finally:
        cur.close(); conn.close()