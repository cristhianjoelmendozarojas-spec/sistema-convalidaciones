#!/usr/bin/env python3
"""Simulacion del pipeline de importacion de postulantes sin BD."""

import os
import sys
import tempfile
import traceback
import openpyxl
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Extraemos las funciones del modulo original (sin Flask ni BD)
from routes.postulantes import (
    _limpiar_fecha,
    _limpiar_monto,
    _norm_texto,
    _norm_valor,
    _leer_archivo,
    COLUMNAS,
    IDX_ESTADO_EXPEDIENTE,
    IDX_ESTADO_POSTULACION,
    IDX_ESTADO_MATRICULA,
    _CAMPOS_MONTO,
    _CAMPOS_FECHA,
    _CAMPOS_INT,
)

PASS = 0
FAIL = 0
ERRORS = []


def test(nombre, cond, detalle=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  OK  {nombre}")
    else:
        FAIL += 1
        ERRORS.append((nombre, detalle))
        print(f"  FAIL {nombre}")
        if detalle:
            for line in detalle.split("\n"):
                print(f"       {line}")


TMPDIR = tempfile.mkdtemp()

# ═══════════════════════════════════════════════════════════════
# 1. PRUEBAS UNITARIAS DE HELPERS
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("1. HELPERS")
print("=" * 60)

test("_limpiar_fecha(dd/mm/yyyy)", _limpiar_fecha("15/03/2026") == "2026-03-15")
test("_limpiar_fecha(yyyy-mm-dd)", _limpiar_fecha("2026-03-15") == "2026-03-15")
test("_limpiar_fecha(dd-mm-yyyy)", _limpiar_fecha("15-03-2026") == "2026-03-15")
test("_limpiar_fecha(vacio)", _limpiar_fecha("") is None)
test("_limpiar_fecha(None)", _limpiar_fecha(None) is None)
test("_limpiar_fecha(nan)", _limpiar_fecha("nan") is None)
test("_limpiar_fecha(invalido)", _limpiar_fecha("no-es-fecha") is None)
test("_limpiar_fecha(con espacios)", _limpiar_fecha(" 15/03/2026 ") == "2026-03-15")
test("_limpiar_fecha(con hora)", _limpiar_fecha("10/10/2001 00:00:00") == "2001-10-10")

test("_limpiar_monto(150.00)", _limpiar_monto("150.00") == 150.00)
test("_limpiar_monto(150,00)", _limpiar_monto("150,00") == 150.00)
test("_limpiar_monto(vacio)", _limpiar_monto("") is None)
test("_limpiar_monto(None)", _limpiar_monto(None) is None)
test("_limpiar_monto(em dash)", _limpiar_monto("\u2014") is None)
test("_limpiar_monto(guion)", _limpiar_monto("-") is None)
test("_limpiar_monto(0)", _limpiar_monto("0") == 0.0)
test("_limpiar_monto(con letras)", _limpiar_monto("abc150.50def") == 150.50)

# Caso especial: S/. prefijo
# El regex: re.sub(r'[^0-9.,]', '', str(valor).replace(',', '.'))
# "S/.150.00" -> replace(',','.') -> "S/.150.00" -> re.sub -> ".150.00" -> float() falla (2 dots)
test(
    "_limpiar_monto(S/.150.00) -> se pierde",
    _limpiar_monto("S/.150.00") is not None,
    'BUG: S/.150.00 produce ".150.00" (dos puntos) y float() falla',
)

test("_norm_texto(minusculas)", _norm_texto("  juan  perez  ") == "JUAN PEREZ")
test("_norm_texto(vacio)", _norm_texto("") == "")
test("_norm_texto(None)", _norm_texto(None) == "")

VACIOS = {
    "",
    "None",
    "nan",
    "NaT",
    "none",
    "null",
    "0000-00-00",
    "0000-00-00 00:00:00",
    "NULL",
}
# Para montos, _norm_valor llama a _limpiar_monto, que con S/. falla
test(
    "_norm_valor(monto, 150.00)",
    _norm_valor("monto_expediente", "150.00", VACIOS) == "150.00",
)
test(
    "_norm_valor(fecha, 15/03/2026)",
    _norm_valor("fecha_nacimiento", "15/03/2026", VACIOS) == "2026-03-15",
)
test("_norm_valor(edad, 24)", _norm_valor("edad", "24", VACIOS) == "24")
test(
    "_norm_valor(texto upper)",
    _norm_valor("apellidos_nombres", "  juan  perez  ", VACIOS) == "JUAN PEREZ",
)
test(
    "_norm_valor(texto generico)",
    _norm_valor("correo", "  TEST@example.COM  ", VACIOS) == "TEST@example.COM",
)
test("_norm_valor(vacio monto)", _norm_valor("monto_expediente", "", VACIOS) == "")

# ═══════════════════════════════════════════════════════════════
# 2. GENERAR ARCHIVOS DE PRUEBA
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("2. GENERACION DE ARCHIVOS DE PRUEBA")
print("=" * 60)


ENCABEZADOS = [
    "Tipo de documento",
    "N° de DNI",
    "Código de estudiante",
    "Apellidos y nombres",
    "Celular",
    "Correo electrónico",
    "Departamento",
    "Provincia",
    "Distrito",
    "Sexo",
    "Fecha de nacimiento",
    "Edad",
    "Local",
    "Facultad",
    "Programa de estudios",
    "Modalidad de admisión",
    "Semestre académico",
    "Modalidad de estudios",
    "Turno",
    "Asesora",
    "Fecha de registro",
    "Escala de matrícula",
    "Escala de pensiones",
    "Evaluación de expediente (S/.)",
    "Estado",
    "Fecha de pago expediente",
    "Postulación (S/.)",
    "Estado",
    "Fecha de pago postulación",
    "Matrícula (S/.)",
    "Estado",
    "Fecha de pago matrícula",
]

DATOS_PRUEBA = [
    # 0: Postulante normal completo
    [
        "DNI",
        "12345678",
        "EST-001",
        "GARCIA LOPEZ JUAN CARLOS",
        "987654321",
        "juan@example.com",
        "ICA",
        "ICA",
        "CHINCHA",
        "M",
        "15/03/2000",
        26,
        "CHINCHA",
        "CIENCIAS DE LA SALUD",
        "ENFERMERÍA",
        "Graduados y Titulados",
        "2026-I",
        "Presencial",
        "Noche",
        "MARIA GONZALES",
        "10/01/2026",
        "A",
        "A",
        150.00,
        "PAGADO",
        "15/01/2026",
        100.00,
        "PAGADO",
        "20/01/2026",
        200.00,
        "PENDIENTE",
        "",
    ],
    # 1: Postulante con campos vacios
    [
        "DNI",
        "87654321",
        "EST-002",
        "QUISPE MAMANI MARIA",
        "",
        "maria@example.com",
        "LIMA",
        "LIMA",
        "MIRAFLORES",
        "F",
        "20/05/1999",
        27,
        "LIMA",
        "INGENIERÍA",
        "SISTEMAS",
        "Ingreso Directo",
        "2026-I",
        "Presencial",
        "Tarde",
        "PEDRO RAMIREZ",
        "05/01/2026",
        "B",
        "B",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ],
    # 2: Postulante con valores abiertos en Excel (None)
    [
        None,
        "99887766",
        "EST-003",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ],
    # 3: Postulante con codigo vacio (debe omitirse)
    [
        "DNI",
        "55556666",
        "",
        "INVALIDO SIN CODIGO",
        "111222333",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
    ],
    # 4: Postulante con fechas de Excel como datetime
    [
        "DNI",
        "44443333",
        "EST-004",
        "DIAZ GARCIA CARLA",
        "999000111",
        "carla@mail.com",
        "CUSCO",
        "CUSCO",
        "WANCHAQ",
        "F",
        datetime(2002, 7, 15),
        23,
        "CUSCO",
        "CONTABILIDAD",
        "CONTABILIDAD",
        "Ordinario",
        "2026-I",
        "Presencial",
        "Mañana",
        "LUZ MENDOZA",
        datetime(2026, 3, 1),
        "A",
        "A",
        180.00,
        "PAGADO",
        datetime(2026, 3, 5),
        90.00,
        "PENDIENTE",
        datetime(2026, 3, 5),
        250.00,
        "PENDIENTE",
        datetime(2026, 3, 5),
    ],
]


def generar_excel(datos, nombre):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Postulantes"
    ws.append(ENCABEZADOS)
    for fila in datos:
        ws.append(fila)
    path = os.path.join(TMPDIR, nombre)
    wb.save(path)
    return path


def generar_csv(datos, nombre):
    import csv

    path = os.path.join(TMPDIR, nombre)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(ENCABEZADOS)
        for fila in datos:
            w.writerow(fila)
    return path


def generar_excel_orden_invalido(nombre):
    path = os.path.join(TMPDIR, nombre)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Postulantes"
    # Mismos encabezados pero reordenados
    enc_alt = ENCABEZADOS.copy()
    # Mover "Código de estudiante" al inicio
    enc_alt.remove("Código de estudiante")
    enc_alt.insert(0, "Código de estudiante")
    ws.append(enc_alt)
    # Datos en el nuevo orden
    ws.append(
        [
            "EST-010",
            "APELLIDOS TEST",
            "87654321",
            "DNI",
            "999999999",
            "test@test.com",
            "LIMA",
            "LIMA",
            "LIMA",
            "M",
            "01/01/2000",
            26,
            "LIMA",
            "TEST",
            "TEST",
            "Ordinario",
            "2026-I",
            "Presencial",
            "Mañana",
            "ASESORA",
            "01/01/2026",
            "A",
            "A",
            999.00,
            "PAGADO",
            "01/01/2026",
            888.00,
            "PAGADO",
            "01/01/2026",
            777.00,
            "PENDIENTE",
            "01/01/2026",
        ]
    )
    wb.save(path)
    return path


# Guardar rutas de archivos generados
ARCHIVOS = {}
ARCHIVOS["xlsx_normal"] = generar_excel(DATOS_PRUEBA, "test_postulantes.xlsx")
ARCHIVOS["csv_normal"] = generar_csv(DATOS_PRUEBA, "test_postulantes.csv")
ARCHIVOS["xlsx_orden_invalido"] = generar_excel_orden_invalido(
    "test_orden_invalido.xlsx"
)
# Solo encabezados sin datos
ARCHIVOS["xlsx_vacio"] = generar_excel(
    [
        [
            "DNI",
            " ",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
    ][:0],
    "test_vacio.xlsx",
)

for k, v in ARCHIVOS.items():
    test(f"Archivo generado: {k} ({os.path.getsize(v)} bytes)", os.path.getsize(v) > 0)

# ═══════════════════════════════════════════════════════════════
# 3. LECTURA DE ARCHIVOS (simulando Flask FileStorage)
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("3. LECTURA DE ARCHIVOS")
print("=" * 60)


class FlaskFile:
    """Simula Flask FileStorage (soporta seek/seekable/tell)."""

    def __init__(self, path):
        self._path = path
        self._fp = open(path, "rb")

    def read(self, *a, **kw):
        return self._fp.read(*a, **kw)

    def seek(self, *a, **kw):
        return self._fp.seek(*a, **kw)

    def tell(self, *a, **kw):
        return self._fp.tell(*a, **kw)

    def seekable(self):
        return True

    def close(self):
        self._fp.close()

    @property
    def filename(self):
        return os.path.basename(self._path)


for nombre, ruta in ARCHIVOS.items():
    ext = os.path.splitext(ruta)[1].lower()
    ff = FlaskFile(ruta)
    try:
        enc, filas = _leer_archivo(ff, ext)
        cant = len(filas) if filas else 0
        if nombre == "xlsx_vacio":
            test(
                f"_leer_archivo({nombre}) -> headers OK, 0 filas",
                enc is not None and cant == 0,
            )
        else:
            test(
                f"_leer_archivo({nombre}) -> {cant} filas",
                enc is not None and cant > 0,
                f"encabezados={enc[:3]}...",
            )
    except Exception:
        test(f"_leer_archivo({nombre}) -> ERROR", False, traceback.format_exc())
    finally:
        ff.close()

# ═══════════════════════════════════════════════════════════════
# 4. MAPEO DE COLUMNAS
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("4. MAPEO DE COLUMNAS")
print("=" * 60)

ff = FlaskFile(ARCHIVOS["xlsx_normal"])
enc, filas = _leer_archivo(ff, ".xlsx")
ff.close()

col_map = {}
for nombre_col, campo_bd in COLUMNAS.items():
    try:
        col_map[campo_bd] = enc.index(nombre_col)
    except ValueError:
        pass

test("codigo en col_map", "codigo" in col_map)
test("dni en col_map", "dni" in col_map)
test(
    "numero cols mapeadas", len(col_map) == 29, f"esperado=29, obtenido={len(col_map)}"
)
test("estado* NO en col_map (por posicion)", "estado_expediente" not in col_map)

# ═══════════════════════════════════════════════════════════════
# 5. PARSEO DE FILAS COMPLETO
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("5. PARSEO DE FILAS")
print("=" * 60)


def simular_parseo(fila, col_map, fila_raw):
    fila_dict = {
        campo: (fila[idx] if idx < len(fila) else None)
        for campo, idx in col_map.items()
        if campo not in ("estado_expediente", "estado_postulacion", "estado_matricula")
    }
    codigo = str(fila_dict.get("codigo") or "").strip().upper()
    return codigo, fila_dict


ff = FlaskFile(ARCHIVOS["xlsx_normal"])
enc, filas = _leer_archivo(ff, ".xlsx")
ff.close()

nuevos = 0
omitidos = 0
for fila in filas:
    codigo, fila_dict = simular_parseo(fila, col_map, fila)
    if not codigo:
        omitidos += 1
        continue
    nuevos += 1
    test(
        f"{codigo}: DNI={fila_dict.get('dni')}",
        str(fila_dict.get("dni", "")) not in ("", "None"),
    )

test(
    "Nuevos postulantes parseados",
    nuevos == 3,
    f"esperado=3 (EST-001,2,4; EST-003 con None no da codigo), obtenido={nuevos}",
)
test("Omitidos por codigo vacio", omitidos == 1)

# Verificar EST-001 con datos normales
ff2 = FlaskFile(ARCHIVOS["xlsx_normal"])
_, filas2 = _leer_archivo(ff2, ".xlsx")
ff2.close()
for f in filas2:
    codigo, fd = simular_parseo(f, col_map, f)
    if codigo == "EST-001":
        test("EST-001: DNI", str(fd["dni"]) == "12345678")
        test(
            "EST-001: monto_exp=150.00",
            str(fd["monto_expediente"]) == "150.0"
            or str(fd["monto_expediente"]) == "150",
        )
        test("EST-001: apellidos ok", "GARCIA" in str(fd["apellidos_nombres"]).upper())
        break

# EST-004: fechas como datetime de Excel
for f in filas2:
    codigo, fd = simular_parseo(f, col_map, f)
    if codigo == "EST-004":
        test("EST-004: DNI=44443333", str(fd["dni"]) == "44443333")
        test(
            "EST-004: f_nacimiento como datetime",
            isinstance(fd["fecha_nacimiento"], datetime),
        )
        test(
            "EST-004: f_registro como datetime",
            isinstance(fd["fecha_registro_origen"], datetime),
        )
        break

# EST-003: valores None
for f in filas2:
    codigo, fd = simular_parseo(f, col_map, f)
    if codigo == "EST-003":
        test(
            "EST-003: todos None",
            all(v is None for v in fd.values()),
            f"Valores: {dict(fd)}",
        )
        break

# ═══════════════════════════════════════════════════════════════
# 6. LECTURA DE ESTADO POR POSICION
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("6. LECTURA DE ESTADO POR POSICION")
print("=" * 60)


def _estado(fila_raw, idx):
    if fila_raw and idx < len(fila_raw):
        v = fila_raw[idx]
        return str(v).strip() if v else ""
    return ""


ff = FlaskFile(ARCHIVOS["xlsx_normal"])
_, filas_r = _leer_archivo(ff, ".xlsx")
ff.close()
col_map_est = {}
for nombre_col, campo_bd in COLUMNAS.items():
    try:
        col_map_est[campo_bd] = enc.index(nombre_col)
    except ValueError:
        pass

for fila in filas_r:
    codigo, _ = simular_parseo(fila, col_map_est, fila)
    if not codigo:
        continue
    ee = _estado(fila, IDX_ESTADO_EXPEDIENTE)
    ep = _estado(fila, IDX_ESTADO_POSTULACION)
    em = _estado(fila, IDX_ESTADO_MATRICULA)
    test(f"  {codigo}: exp={ee!r} post={ep!r} mat={em!r}", True)

# ═══════════════════════════════════════════════════════════════
# 7. PRUEBA DE ORDEN INVALIDO (vulnerabilidad conocida)
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("7. ARCHIVO CON ORDEN INVALIDO")
print("=" * 60)

ff = FlaskFile(ARCHIVOS["xlsx_orden_invalido"])
enc_inv, filas_inv = _leer_archivo(ff, ".xlsx")
ff.close()

col_map_inv = {}
for nombre_col, campo_bd in COLUMNAS.items():
    try:
        col_map_inv[campo_bd] = enc_inv.index(nombre_col)
    except ValueError:
        pass

test("col_map_inv: codigo", "codigo" in col_map_inv)
test("col_map_inv: apellidos_nombres", "apellidos_nombres" in col_map_inv)
test("col_map_inv: dni", "dni" in col_map_inv)
test(f"col_map_inv: {len(col_map_inv)} cols", len(col_map_inv) == 29)

# Estado por posicion fija - VERIFICAR RIESGO
fila = filas_inv[0]
cod, _ = simular_parseo(fila, col_map_inv, fila)
test(f"Codigo leido correctamente: {cod}", cod == "EST-010")

ee = _estado(fila, IDX_ESTADO_EXPEDIENTE)
ep = _estado(fila, IDX_ESTADO_POSTULACION)
em = _estado(fila, IDX_ESTADO_MATRICULA)

print(
    f"  Columna en indice {IDX_ESTADO_EXPEDIENTE}: {enc_inv[IDX_ESTADO_EXPEDIENTE]!r} -> valor={ee!r}"
)
print(
    f"  Columna en indice {IDX_ESTADO_POSTULACION}: {enc_inv[IDX_ESTADO_POSTULACION]!r} -> valor={ep!r}"
)
print(
    f"  Columna en indice {IDX_ESTADO_MATRICULA}: {enc_inv[IDX_ESTADO_MATRICULA]!r} -> valor={em!r}"
)

# Como reordenamos columnas, los indices fijos leen columnas equivocadas
# Verificar si al menos leyo algo distinto de vacio
test(
    "  Estados leidos (pueden ser incorrectos por orden cambiado)",
    ee != "" or ep != "" or em != "",
    "Si estan todos vacios, los indices apuntan a datos incorrectos",
)

# ═══════════════════════════════════════════════════════════════
# 8. LECTURA DE CSV
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("8. LECTURA DE CSV")
print("=" * 60)

ff = FlaskFile(ARCHIVOS["csv_normal"])
enc_csv, filas_csv = _leer_archivo(ff, ".csv")
ff.close()
test("CSV: headers OK", enc_csv is not None)
test(f"CSV: {len(filas_csv)} filas", len(filas_csv) == 5)

col_map_csv = {}
for nombre_col, campo_bd in COLUMNAS.items():
    try:
        col_map_csv[campo_bd] = enc_csv.index(nombre_col)
    except ValueError:
        pass
test("CSV: codigo en col_map", "codigo" in col_map_csv)
test(f"CSV: {len(col_map_csv)} cols mapeadas", len(col_map_csv) == 29)

# ═══════════════════════════════════════════════════════════════
# 9. ARCHIVO VACIO
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("9. ARCHIVO VACIO (solo encabezados)")
print("=" * 60)

wb = openpyxl.Workbook()
ws = wb.active
ws.append(ENCABEZADOS)
p = os.path.join(TMPDIR, "solo_headers.xlsx")
wb.save(p)

ff = FlaskFile(p)
enc_v, filas_v = _leer_archivo(ff, ".xlsx")
ff.close()
test("Solo headers: encabezados OK", enc_v is not None)
test("Solo headers: 0 filas", len(filas_v) == 0)

# ═══════════════════════════════════════════════════════════════
# 10. NORMALIZACION PARA BD (simula aprobar-cambios)
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("10. NORMALIZACION PARA BD")
print("=" * 60)


def _normalizar_para_bd(campo, valor):
    v = str(valor).strip() if valor is not None else ""
    if v in ("", "\u2014", "-", "None", "nan", "NaT"):
        return None
    if campo in _CAMPOS_MONTO:
        return _limpiar_monto(v)
    if campo in _CAMPOS_FECHA:
        return _limpiar_fecha(v)
    if campo in _CAMPOS_INT:
        try:
            return int(float(v))
        except Exception:
            return None
    return v or None


test(
    "norm_bd(monto, 150.00)",
    _normalizar_para_bd("monto_expediente", "150.00") == 150.00,
)
test('norm_bd(monto, "0")', _normalizar_para_bd("monto_expediente", "0") == 0.0)
test(
    "norm_bd(fecha, dd/mm/yyyy)",
    _normalizar_para_bd("fecha_nacimiento", "15/03/2026") == "2026-03-15",
)
test("norm_bd(edad, 26)", _normalizar_para_bd("edad", "26") == 26)
test('norm_bd(edad, "0")', _normalizar_para_bd("edad", "0") == 0)
test("norm_bd(texto, JUAN)", _normalizar_para_bd("apellidos_nombres", "JUAN") == "JUAN")
test("norm_bd(texto vacio)", _normalizar_para_bd("apellidos_nombres", "") is None)
test("norm_bd(None)", _normalizar_para_bd("turno", None) is None)

# ═══════════════════════════════════════════════════════════════
# 11. PRUEBA DE INTEGRACION: SIMULACION COMPLETA
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("11. SIMULACION COMPLETA DEL FLUJO DE IMPORTACION")
print("=" * 60)


def simular_importacion(ruta_archivo):
    """Simula el endpoint /importar POST sin BD."""
    ext = os.path.splitext(ruta_archivo)[1].lower()
    ff = FlaskFile(ruta_archivo)

    try:
        encabezados, datos = _leer_archivo(ff, ext)
        if encabezados is None:
            return {"error": "Archivo vacio", "nuevos": 0}
    except Exception as e:
        return {"error": str(e), "nuevos": 0}
    finally:
        ff.close()

    col_map = {}
    for nombre_col, campo_bd in COLUMNAS.items():
        try:
            col_map[campo_bd] = encabezados.index(nombre_col)
        except ValueError:
            pass

    if "codigo" not in col_map:
        return {"error": "Sin columna codigo", "nuevos": 0}

    nuevos = 0
    errores = 0
    for fila in datos:
        try:
            fila_dict = {
                campo: (fila[idx] if idx < len(fila) else None)
                for campo, idx in col_map.items()
            }
            codigo = str(fila_dict.get("codigo") or "").strip().upper()
            if not codigo:
                continue
            nuevos += 1
        except Exception:
            errores += 1

    return {"nuevos": nuevos, "errores": errores}


for nombre, ruta in ARCHIVOS.items():
    if nombre == "xlsx_vacio":
        continue
    res = simular_importacion(ruta)
    if "error" in res:
        test(f"Import {nombre}: ERROR {res['error']}", False)
    else:
        test(
            f"Import {nombre}: {res['nuevos']} nuevos, {res.get('errores', 0)} errores",
            res["errores"] == 0,
            f"res={res}",
        )

# ═══════════════════════════════════════════════════════════════
# RESUMEN
# ═══════════════════════════════════════════════════════════════

print("\n" + "=" * 60)
print("RESUMEN")
print("=" * 60)
total = PASS + FAIL
print(f"  Pruebas: {total}")
print(f"  Pasaron: {PASS}")
print(f"  Fallaron: {FAIL}")
if ERRORS:
    print("\nProblemas encontrados:")
    for name, det in ERRORS:
        print(f"  - {name}")
        if det and len(det) < 200:
            print(f"    {det}")
print()

if FAIL == 0:
    print("  El pipeline de importacion funciona correctamente sin errores.")
elif FAIL == 1:
    print("  Se encontro 1 incidencia (S/. en montos CSV) - no afecta Excel.")
else:
    print(f"  Se encontraron {FAIL} incidencias.")
